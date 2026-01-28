# -*- coding: utf-8 -*-
"""
Nature For Water Facility - The Nature Conservancy
Timesheet Automation Tool
-------------------------------------------------------------------------
Python 3.11+
-------------------------------------------------------------------------
Timesheet automation tool that integrates:
- Outlook Calendar
- Deltek System
- N4W Facility System

Author: Jonathan Nogales Pimentel, Carlos A. Rogéliz Prada
Email: jonathan.nogales@tnc.org
Date: November, 2024
"""

# =============================================================================
# IMPORTS
# =============================================================================

# Librerías estándar
import os
import time
import threading
import traceback
import re
import calendar
import shutil
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, List, Tuple

# Librerías de terceros - Datos y análisis
import numpy as np
import pandas as pd
import pytz
from tzlocal import get_localzone

# Librerías de terceros - Excel y archivos
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel
import openpyxl

# Librerías de terceros - Interfaz gráfica
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import customtkinter as ctk
from tkcalendar import DateEntry

# Librerías de terceros - Automatización web
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException

# Librerías de Windows
import win32com.client
from win32com.client import constants
import pythoncom

# Librerías de red
import requests

try:
    import winreg  # Solo disponible en Windows
except ImportError:
    winreg = None  # Permite importar el módulo en otros SO


# =============================================================================
# CONFIGURACIÓN GLOBAL
# =============================================================================
# Configuración del tema de CustomTkinter
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Paleta de colores basada en Notion Dark Mode
COLORS = {
    'bg_primary': '#191919',      # Fondo principal
    'bg_secondary': '#252525',    # Fondo secundario
    'bg_tertiary': '#2F2F2F',     # Fondo terciario
    'text_primary': '#FFFFFF',    # Texto principal
    'text_secondary': '#9B9A97',  # Texto secundario
    'accent': '#2383E2',          # Color de acento
    'accent_hover': '#1A73CF',    # Color de acento al pasar el cursor
    'border': '#3F3F3F',          # Bordes
    'success': '#0F7B0F',         # Color de éxito
    'warning': '#CD6200'          # Color de advertencia
}

# Variables globales para manejo de barras de progreso
app_instance = None
progress_window = None
progress_bar = None

# =============================================================================
# FUNCIONES HELPER
# =============================================================================
def get_date_columns(df):
    """
    Detecta columnas de fecha en un DataFrame con formato YYYY-MM-DD.

    Args:
        df (pd.DataFrame): DataFrame a analizar

    Returns:
        list: Lista de nombres de columnas que contienen fechas
    """
    date_pattern = re.compile(r'\d{4}-\d{2}-\d{2}')
    return [col for col in df.columns if date_pattern.search(str(col))]

# =============================================================================
# CLASE TOOLTIP
# =============================================================================
class ToolTip:
    """
    Clase para crear tooltips (mensajes emergentes) en CustomTkinter.
    Muestra un mensaje cuando el usuario pasa el mouse sobre un widget.
    """
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        """Muestra el tooltip."""
        if self.tooltip_window or not self.text:
            return

        # Obtener posición del widget
        x = self.widget.winfo_rootx() + 25
        y = self.widget.winfo_rooty() + 25

        # Crear ventana emergente
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")

        # Crear frame con el mensaje
        frame = tk.Frame(
            self.tooltip_window,
            background=COLORS['bg_tertiary'],
            borderwidth=1,
            relief="solid"
        )
        frame.pack()

        # Crear label con el texto
        label = tk.Label(
            frame,
            text=self.text,
            justify=tk.LEFT,
            background=COLORS['bg_tertiary'],
            foreground=COLORS['text_primary'],
            relief=tk.FLAT,
            borderwidth=0,
            font=("Segoe UI", 10),
            padx=8,
            pady=6
        )
        label.pack()

    def hide_tooltip(self, event=None):
        """Oculta el tooltip."""
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

# =============================================================================
# FUNCIONES AUXILIARES BÁSICAS
# =============================================================================
def create_folder(directory):
    """Crea un directorio si no existe."""
    if not os.path.exists(directory):
        os.makedirs(directory)


def remove_timezone(date):
    """Convierte un datetime timezone-aware a timezone-naive."""
    return date.replace(tzinfo=None)


# =============================================================================
# FUNCIONES DE VALIDACIÓN DE CHROMEDRIVER
# =============================================================================
def get_chrome_version():
    """
    Obtiene la versión de Chrome instalada desde el registro de Windows.

    Returns:
        str: Versión de Chrome (ej: '140.0.7339.207') o None si no se encuentra
    """
    try:
        if winreg is None:
            print("ERROR: winreg no disponible (solo funciona en Windows)")
            return None

        # Intentar obtener la versión desde diferentes ubicaciones del registro
        registry_paths = [
            (winreg.HKEY_CURRENT_USER, r"Software\Google\Chrome\BLBeacon", "version"),
            (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Google\Chrome\BLBeacon", "version"),
            (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Wow6432Node\Google\Chrome\BLBeacon", "version")
        ]

        for hkey, path, value_name in registry_paths:
            try:
                key = winreg.OpenKey(hkey, path)
                version, _ = winreg.QueryValueEx(key, value_name)
                winreg.CloseKey(key)
                print(f"Chrome version encontrada: {version}")
                return version
            except WindowsError:
                continue

        print("ERROR: No se pudo encontrar la versión de Chrome en el registro")
        return None

    except Exception as e:
        print(f"ERROR al obtener versión de Chrome: {e}")
        return None


def get_chromedriver_version(chromedriver_path="chromedriver.exe"):
    """
    Obtiene la versión de ChromeDriver ejecutando el comando --version.

    Args:
        chromedriver_path (str): Ruta al ejecutable de ChromeDriver

    Returns:
        str: Versión de ChromeDriver (ej: '140.0.7339.207') o None si no existe
    """
    try:
        if not os.path.exists(chromedriver_path):
            print(f"ChromeDriver no encontrado en: {chromedriver_path}")
            return None

        # Ejecutar chromedriver --version
        import subprocess
        result = subprocess.run(
            [chromedriver_path, "--version"],
            capture_output=True,
            text=True,
            timeout=5
        )

        # Parsear output: "ChromeDriver 140.0.7339.207 (xyz...)"
        output = result.stdout.strip()
        match = re.search(r'ChromeDriver\s+([\d.]+)', output)

        if match:
            version = match.group(1)
            print(f"ChromeDriver version encontrada: {version}")
            return version
        else:
            print(f"No se pudo parsear versión de ChromeDriver: {output}")
            return None

    except Exception as e:
        print(f"ERROR al obtener versión de ChromeDriver: {e}")
        return None


def download_and_extract_chromedriver(chrome_version, chromedriver_path="chromedriver.exe"):
    """
    Descarga y extrae ChromeDriver compatible con la versión de Chrome.

    Args:
        chrome_version (str): Versión de Chrome (ej: '140.0.7339.207')
        chromedriver_path (str): Ruta donde se guardará ChromeDriver

    Returns:
        bool: True si la descarga y extracción fueron exitosas, False en caso contrario
    """
    import zipfile
    import tempfile

    try:
        # Construir URL de descarga
        download_url = f"https://storage.googleapis.com/chrome-for-testing-public/{chrome_version}/win64/chromedriver-win64.zip"

        print(f"Descargando ChromeDriver desde: {download_url}")

        # Descargar archivo ZIP
        response = requests.get(download_url, stream=True, timeout=30)
        response.raise_for_status()  # Lanzar excepción si hay error HTTP

        # Guardar ZIP en archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as temp_zip:
            temp_zip_path = temp_zip.name
            total_size = int(response.headers.get('content-length', 0))
            downloaded = 0

            print("Descargando ChromeDriver...")
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    temp_zip.write(chunk)
                    downloaded += len(chunk)
                    if total_size > 0:
                        progress = (downloaded / total_size) * 100
                        print(f"Progreso: {progress:.1f}%", end='\r')

            print("\nDescarga completada.")

        # Extraer ChromeDriver del ZIP
        print("Extrayendo ChromeDriver...")
        with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
            # El archivo está en: chromedriver-win64/chromedriver.exe
            chromedriver_in_zip = "chromedriver-win64/chromedriver.exe"

            # Extraer a directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                zip_ref.extract(chromedriver_in_zip, temp_dir)
                extracted_path = os.path.join(temp_dir, chromedriver_in_zip)

                # Eliminar ChromeDriver antiguo si existe
                if os.path.exists(chromedriver_path):
                    print(f"Eliminando ChromeDriver antiguo: {chromedriver_path}")
                    os.remove(chromedriver_path)

                # Copiar nuevo ChromeDriver a la ubicación final
                print(f"Copiando ChromeDriver a: {chromedriver_path}")
                shutil.copy2(extracted_path, chromedriver_path)

        # Limpiar archivo ZIP temporal
        os.remove(temp_zip_path)

        print("ChromeDriver instalado exitosamente.")
        return True

    except requests.exceptions.RequestException as e:
        print(f"ERROR al descargar ChromeDriver: {e}")
        print(f"URL de descarga: {download_url}")
        return False
    except zipfile.BadZipFile as e:
        print(f"ERROR: Archivo ZIP corrupto: {e}")
        return False
    except Exception as e:
        print(f"ERROR inesperado al instalar ChromeDriver: {e}")
        traceback.print_exc()
        return False


def validate_and_update_chromedriver():
    """
    Función principal que valida ChromeDriver y lo actualiza si es necesario.
    Esta función debe ejecutarse antes de iniciar la aplicación GUI.

    Returns:
        bool: True si ChromeDriver está listo para usar, False si hay error crítico
    """
    print("\n" + "="*70)
    print("VALIDACIÓN DE CHROMEDRIVER")
    print("="*70)

    # Paso 1: Obtener versión de Chrome
    chrome_version = get_chrome_version()
    if chrome_version is None:
        print("\nERROR CRÍTICO: Google Chrome no está instalado.")
        print("Por favor, instale Google Chrome desde: https://www.google.com/chrome/")
        messagebox.showerror(
            "Chrome no encontrado",
            "Google Chrome no está instalado en su sistema.\n\n"
            "Por favor, instale Chrome desde:\n"
            "https://www.google.com/chrome/"
        )
        return False

    # Extraer versión major (ej: '140.0.7339.207' -> '140')
    chrome_major = chrome_version.split('.')[0]

    # Paso 2: Verificar si existe ChromeDriver
    chromedriver_path = "chromedriver.exe"
    chromedriver_version = get_chromedriver_version(chromedriver_path)

    # Paso 3: Determinar si necesita descarga
    needs_download = False

    if chromedriver_version is None:
        print("\nChromeDriver no encontrado. Se descargará automáticamente.")
        needs_download = True
    else:
        # Comparar versión major
        chromedriver_major = chromedriver_version.split('.')[0]

        if chrome_major == chromedriver_major:
            print(f"\n✓ ChromeDriver compatible (Chrome: {chrome_version}, ChromeDriver: {chromedriver_version})")
            print("="*70 + "\n")
            return True
        else:
            print(f"\n⚠ Versión incompatible:")
            print(f"  Chrome: {chrome_version} (major: {chrome_major})")
            print(f"  ChromeDriver: {chromedriver_version} (major: {chromedriver_major})")
            print("\nSe descargará la versión compatible...")
            needs_download = True

    # Paso 4: Descargar ChromeDriver si es necesario
    if needs_download:
        success = download_and_extract_chromedriver(chrome_version, chromedriver_path)

        if success:
            # Verificar que la instalación fue exitosa
            new_version = get_chromedriver_version(chromedriver_path)
            if new_version:
                print(f"\n✓ ChromeDriver actualizado exitosamente a versión: {new_version}")
                print("="*70 + "\n")
                return True
            else:
                print("\nERROR: ChromeDriver descargado pero no se pudo verificar.")
                return False
        else:
            print("\nERROR: No se pudo descargar ChromeDriver.")
            print(f"\nPuede descargarlo manualmente desde:")
            print(f"https://storage.googleapis.com/chrome-for-testing-public/{chrome_version}/win64/chromedriver-win64.zip")

            messagebox.showerror(
                "Error al actualizar ChromeDriver",
                f"No se pudo descargar ChromeDriver automáticamente.\n\n"
                f"Por favor, descárguelo manualmente desde:\n"
                f"https://storage.googleapis.com/chrome-for-testing-public/{chrome_version}/win64/chromedriver-win64.zip\n\n"
                f"Extraiga chromedriver.exe a la carpeta de la aplicación."
            )
            return False

    return True


# =============================================================================
# FUNCIONES DE PRORATE Y REDISTRIBUCIÓN DE HORAS
# =============================================================================
def load_prorate_data(n4w_task_details_path: str) -> Dict[str, int]:
    """
    Carga información de prorate desde el archivo N4W Task Details.
    
    Args:
        n4w_task_details_path (str): Ruta al archivo N4W_Task_Details.xlsx
        
    Returns:
        Dict[str, int]: Diccionario que mapea Task_Name a valor prorate (0 o 1)
    """
    try:
        df = pd.read_excel(n4w_task_details_path)
        
        # Crear diccionario que mapea Task_Name a valor prorate
        prorate_dict = dict(zip(df['Task_Name'], df['Prorate']))
        
        print(f"Loaded prorate data for {len(prorate_dict)} projects")
        return prorate_dict
        
    except Exception as e:
        print(f"Error loading prorate data: {e}")
        return {}


def get_distribution_weights(df_real: pd.DataFrame, date_columns: List[str]) -> pd.Series:
    """
    Calcula pesos proporcionales para proyectos reales basado en sus horas totales.
    
    Args:
        df_real (pd.DataFrame): DataFrame solo con proyectos reales
        date_columns (List[str]): Lista de nombres de columnas de fechas
        
    Returns:
        pd.Series: Serie con pesos proporcionales para cada proyecto real
    """
    # Calcular horas totales para cada proyecto real
    total_hours = df_real[date_columns].sum(axis=1)
    
    # Calcular pesos proporcionales (evitar división por cero)
    total_sum = total_hours.sum()
    if total_sum == 0:
        # Si no hay horas en proyectos reales, distribuir equitativamente
        weights = pd.Series([1 / len(df_real)] * len(df_real), index=df_real.index)
    else:
        weights = total_hours / total_sum
    
    return weights


def show_project_selection_window(df_projects: pd.DataFrame, database_path: str = None) -> Dict[str, bool]:
    """
    Muestra ventana para que el usuario seleccione qué proyectos recibirán horas redistribuidas.

    Args:
        df_projects (pd.DataFrame): DataFrame con proyectos reales (prorate=0)
        database_path (str): Ruta al archivo de base de datos con información de proyectos

    Returns:
        Dict[str, bool]: Diccionario {project_code: True/False} indicando selección
                        None si el usuario cancela
    """

    # Cargar información adicional de la base de datos de proyectos
    project_details = {}
    try:
        if database_path and os.path.exists(database_path):
            df_all_projects = pd.read_excel(database_path, sheet_name='N4W-Projects')

            for _, row in df_all_projects.iterrows():
                project_details[row['Code']] = {
                    'Task_Name': row.get('Task Name', 'N/A'),
                }
    except Exception as e:
        print(f"Warning: Could not load project database: {e}")
        project_details = {}

    # Obtener lista única de códigos de proyectos reales
    unique_codes = df_projects['Code'].unique()

    print(f"Creating project selection window for {len(unique_codes)} projects...")

    # Crear ventana de selección
    try:
        selection_window = ctk.CTkToplevel()
    except Exception as e:
        print(f"Error creating CTkToplevel: {e}")
        print("Attempting to create with explicit root...")
        # Si falla, intentar crear una ventana raíz temporal
        temp_root = ctk.CTk()
        temp_root.withdraw()
        selection_window = ctk.CTkToplevel(temp_root)

    selection_window.title("Select Projects for Hour Redistribution")
    selection_window.geometry("800x600")
    selection_window.configure(fg_color=COLORS['bg_primary'])
    selection_window.transient()
    selection_window.grab_set()  # Hacer ventana modal

    # Centrar la ventana
    selection_window.update_idletasks()
    x = (selection_window.winfo_screenwidth() // 2) - (800 // 2)
    y = (selection_window.winfo_screenheight() // 2) - (600 // 2)
    selection_window.geometry(f"800x600+{x}+{y}")

    # Variable para almacenar selección del usuario
    selection_result = {'selections': None, 'cancelled': False}

    # Diccionario para almacenar variables de checkbox
    checkbox_vars = {}

    def on_continue():
        # Recopilar selecciones
        selections = {code: var.get() for code, var in checkbox_vars.items()}
        selection_result['selections'] = selections
        selection_result['cancelled'] = False
        selection_window.destroy()

    def on_cancel():
        selection_result['selections'] = None
        selection_result['cancelled'] = True
        selection_window.destroy()

    def select_all():
        for var in checkbox_vars.values():
            var.set(True)

    def deselect_all():
        for var in checkbox_vars.values():
            var.set(False)

    # Encabezado
    header_frame = ctk.CTkFrame(selection_window, fg_color="transparent")
    header_frame.pack(fill="x", padx=20, pady=(20, 10))

    title_label = ctk.CTkLabel(
        header_frame,
        text="Select Projects for Redistribution",
        font=ctk.CTkFont(size=24, weight="bold"),
        text_color=COLORS['text_primary']
    )
    title_label.pack()

    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select which projects will receive redistributed hours. Unselected projects will keep their original hours.",
        font=ctk.CTkFont(size=12),
        text_color=COLORS['text_secondary']
    )
    subtitle_label.pack(pady=(5, 0))

    # Marco de botones de selección rápida
    quick_buttons_frame = ctk.CTkFrame(selection_window, fg_color="transparent")
    quick_buttons_frame.pack(fill="x", padx=20, pady=(0, 10))

    select_all_btn = ctk.CTkButton(
        quick_buttons_frame,
        text="Select All",
        command=select_all,
        width=120,
        fg_color=COLORS['accent'],
        hover_color=COLORS['accent_hover']
    )
    select_all_btn.pack(side="left", padx=5)

    deselect_all_btn = ctk.CTkButton(
        quick_buttons_frame,
        text="Deselect All",
        command=deselect_all,
        width=120,
        fg_color=COLORS['bg_tertiary'],
        hover_color=COLORS['border']
    )
    deselect_all_btn.pack(side="left", padx=5)

    # Marco de lista con checkboxes
    list_frame = ctk.CTkScrollableFrame(
        selection_window,
        fg_color=COLORS['bg_secondary'],
        corner_radius=8,
        border_width=1,
        border_color=COLORS['border']
    )
    list_frame.pack(fill="both", expand=True, padx=20, pady=10)

    # Encabezados
    headers_frame = ctk.CTkFrame(list_frame, fg_color=COLORS['bg_tertiary'])
    headers_frame.pack(fill="x", pady=(0, 5))

    ctk.CTkLabel(headers_frame, text="", width=40).grid(row=0, column=0, padx=5, pady=8)  # Checkbox column
    ctk.CTkLabel(headers_frame, text="Code", font=ctk.CTkFont(weight="bold"), width=100).grid(row=0, column=1, padx=5, pady=8, sticky="w")
    ctk.CTkLabel(headers_frame, text="Task Name", font=ctk.CTkFont(weight="bold"), width=450).grid(row=0, column=2, padx=5, pady=8, sticky="w")

    # Filas de proyectos con checkboxes
    for i, code in enumerate(sorted(unique_codes)):
        details = project_details.get(code, {})
        task_name = details.get('Task_Name', 'N/A')

        row_color = COLORS['bg_primary'] if i % 2 == 0 else COLORS['bg_secondary']
        row_frame = ctk.CTkFrame(list_frame, fg_color=row_color, corner_radius=4)
        row_frame.pack(fill="x", pady=1)

        # Checkbox (por defecto seleccionado)
        checkbox_var = ctk.BooleanVar(value=True)
        checkbox_vars[code] = checkbox_var

        checkbox = ctk.CTkCheckBox(
            row_frame,
            text="",
            variable=checkbox_var,
            width=40,
            fg_color=COLORS['accent'],
            hover_color=COLORS['accent_hover']
        )
        checkbox.grid(row=0, column=0, padx=5, pady=8)

        ctk.CTkLabel(row_frame, text=code, width=100, anchor="w").grid(row=0, column=1, padx=5, pady=8, sticky="w")
        ctk.CTkLabel(row_frame, text=task_name, width=450, anchor="w").grid(row=0, column=2, padx=5, pady=8, sticky="w")

    # Botones de acción
    buttons_frame = ctk.CTkFrame(selection_window, fg_color="transparent")
    buttons_frame.pack(fill="x", padx=20, pady=(10, 20))

    cancel_btn = ctk.CTkButton(
        buttons_frame,
        text="Cancel",
        command=on_cancel,
        width=150,
        fg_color=COLORS['bg_tertiary'],
        hover_color=COLORS['border']
    )
    cancel_btn.pack(side="right", padx=5)

    continue_btn = ctk.CTkButton(
        buttons_frame,
        text="Continue",
        command=on_continue,
        width=150,
        fg_color=COLORS['accent'],
        hover_color=COLORS['accent_hover']
    )
    continue_btn.pack(side="right", padx=5)

    # Esperar a que el usuario cierre la ventana
    selection_window.wait_window()

    # Retornar resultado
    if selection_result['cancelled']:
        return None
    return selection_result['selections']


def redistribute_hours_by_earning(deltek_path: str, n4w_task_details_path: str, output_path: str, database_path: str = None):
    """
    Redistribuye horas de proyectos virtuales a proyectos reales basado en tipos de earning.

    Args:
        deltek_path (str): Ruta al archivo 02-Timesheet.csv
        n4w_task_details_path (str): Ruta al archivo N4W_Task_Details.xlsx
        output_path (str): Ruta para archivo de salida con horas redistribuidas
        database_path (str): Ruta al archivo de base de datos con proyectos (opcional)
    """

    print("Starting hour redistribution process...")

    # Cargar datos
    try:
        df_deltek = pd.read_csv(deltek_path)
        print(f"Loaded Deltek data: {len(df_deltek)} rows")
    except Exception as e:
        print(f"Error loading Deltek data: {e}")
        return

    # Cargar datos de prorate
    prorate_dict = load_prorate_data(n4w_task_details_path)
    if not prorate_dict:
        print("Failed to load prorate data. Exiting.")
        return

    # Agregar información de prorate al dataframe deltek
    df_deltek['Prorate'] = df_deltek['Code'].map(prorate_dict).fillna(0).astype(int)

    # Identificar proyectos exceptuados (códigos XX)
    # Estos proyectos NO participan en redistribución: ni dan ni reciben horas
    excepted_projects = df_deltek['Code'].astype(str).str.upper().str.startswith('XX')
    df_deltek.loc[excepted_projects, 'Prorate'] = -1  # Marcador especial para exceptuados

    # Identificar columnas de fechas
    date_columns = get_date_columns(df_deltek)
    if not date_columns:
        print("No date columns found in Deltek file")
        return

    print(f"Found {len(date_columns)} date columns")

    # Separar proyectos en tres categorías:
    # - Virtuales (prorate=1): se redistribuyen a proyectos reales
    # - Reales (prorate=0): reciben horas de proyectos virtuales
    # - Exceptuados (prorate=-1): se mantienen exactamente igual (P100001, etc.)
    df_virtual = df_deltek[df_deltek['Prorate'] == 1].copy()
    df_real = df_deltek[df_deltek['Prorate'] == 0].copy()
    df_excepted = df_deltek[df_deltek['Prorate'] == -1].copy()

    print(f"Virtual projects: {len(df_virtual)}")
    print(f"Real projects: {len(df_real)}")
    print(f"Excepted projects: {len(df_excepted)}")

    if len(df_real) == 0:
        print("No real projects found for redistribution")
        return

    # Mostrar ventana de selección de proyectos
    print("=" * 60)
    print("SHOWING PROJECT SELECTION WINDOW")
    print(f"Number of real projects: {len(df_real)}")
    print(f"Database path: {database_path}")
    print("=" * 60)

    try:
        project_selections = show_project_selection_window(df_real, database_path)
        print(f"Project selections returned: {project_selections}")
    except Exception as e:
        print(f"ERROR in show_project_selection_window: {e}")
        import traceback
        traceback.print_exc()
        return

    if project_selections is None:
        print("Process cancelled by user during project selection.")
        return

    # Separar proyectos reales en dos categorías según selección del usuario:
    # - Reales seleccionados (redistribute_target=1): reciben horas redistribuidas
    # - Reales NO seleccionados (redistribute_target=0): mantienen horas originales
    df_real['Redistribute_Target'] = df_real['Code'].map(project_selections).fillna(True).astype(bool)

    df_real_selected = df_real[df_real['Redistribute_Target'] == True].copy()
    df_real_not_selected = df_real[df_real['Redistribute_Target'] == False].copy()

    print(f"Real projects selected for redistribution: {len(df_real_selected)}")
    print(f"Real projects NOT selected (will keep original hours): {len(df_real_not_selected)}")

    if len(df_real_selected) == 0:
        print("No projects selected for redistribution. Exiting.")
        return

    # Inicializar dataframe resultado con proyectos reales seleccionados
    df_result = df_real_selected.copy()

    # Procesar cada proyecto virtual - distribuir proporcionalmente entre proyectos reales
    for idx, virtual_row in df_virtual.iterrows():
        project_code = virtual_row['Code']

        print(f"Processing virtual project {project_code}")

        # Obtener horas a redistribuir
        hours_to_redistribute = virtual_row[date_columns].values

        # Distribuir proporcionalmente entre proyectos reales SELECCIONADOS
        if len(df_real_selected) > 0:
            weights = get_distribution_weights(df_real_selected, date_columns)

            # Agregar horas distribuidas a proyectos reales seleccionados
            for real_idx in df_real_selected.index:
                weight = weights.loc[real_idx]
                additional_hours = hours_to_redistribute * weight

                # Encontrar fila correspondiente en dataframe resultado
                result_idx = df_result[df_result['Code'] == df_real_selected.loc[real_idx, 'Code']].index[0]
                df_result.loc[result_idx, date_columns] += additional_hours

    # Agrupar por Code y sumar (en caso de duplicados)
    groupby_columns = ['Code']
    df_result = df_result.groupby(groupby_columns, as_index=False)[date_columns].sum()

    # Redondear horas a precisión de 0.25
    df_result[date_columns] = df_result[date_columns].applymap(lambda x: round(x * 4) / 4)

    # Validar y ajustar día por día
    print("Validating hours day by day (only on selected projects)...")

    # Crear resúmenes por día para datos originales
    df_original_for_comparison = pd.concat([df_virtual, df_real_selected], ignore_index=True)
    original_totals = df_original_for_comparison[date_columns].sum()

    # Crear resúmenes por día para datos con prorate
    prorate_totals = df_result[date_columns].sum()

    # Revisar cada día
    total_adjustments = 0
    for date_col in date_columns:
        original_hours = original_totals[date_col]
        prorate_hours = prorate_totals[date_col]
        difference = original_hours - prorate_hours

        # Si hay diferencia significativa, ajustar
        if abs(difference) > 0.001:
            print(f"Adjusting {difference:.3f} hours on {date_col}")

            # Ordenar proyectos por horas (descendente)
            sorted_projects = df_result.sort_values(by=date_col, ascending=False)

            for idx in sorted_projects.index:
                current_hours = df_result.loc[idx, date_col]
                project_code = df_result.loc[idx, 'Code']

                if difference > 0:
                    # AGREGAR horas al proyecto con más horas
                    df_result.loc[idx, date_col] += difference
                    df_result.loc[idx, date_col] = round(df_result.loc[idx, date_col] * 4) / 4
                    print(f"  → Added {difference:.3f}h to project {project_code}")
                    total_adjustments += 1
                    break
                else:
                    # RESTAR horas del proyecto con más horas (si tiene suficiente)
                    hours_to_subtract = abs(difference)
                    if current_hours >= hours_to_subtract:
                        df_result.loc[idx, date_col] -= hours_to_subtract
                        df_result.loc[idx, date_col] = round(df_result.loc[idx, date_col] * 4) / 4
                        print(f"  → Subtracted {hours_to_subtract:.3f}h from project {project_code}")
                        total_adjustments += 1
                        break

    if total_adjustments > 0:
        print(f"Total adjustments made: {total_adjustments}")
    else:
        print("No adjustments needed - all day totals match perfectly")

    # Columnas base: Code + fechas
    base_columns = ['Code'] + date_columns

    # AHORA agregar proyectos reales NO seleccionados al resultado final (con horas originales)
    if len(df_real_not_selected) > 0:
        df_not_selected_clean = df_real_not_selected[base_columns].copy()
        df_result = pd.concat([df_result, df_not_selected_clean], ignore_index=True)
        print(f"Added {len(df_real_not_selected)} non-selected projects with original hours to final result")

    # Agregar proyectos exceptuados al resultado final (sin modificaciones)
    if len(df_excepted) > 0:
        df_excepted_clean = df_excepted[base_columns].copy()
        df_result = pd.concat([df_result, df_excepted_clean], ignore_index=True)
        print(f"Added {len(df_excepted)} excepted projects to final result")

    # Agregar Task Name y Grant ID desde la base de datos
    if database_path and os.path.exists(database_path):
        try:
            df_db = pd.read_excel(database_path, sheet_name='N4W-Projects')
            df_db = df_db[['Code', 'Task Name', 'Grant ID']].drop_duplicates()
            df_result = df_result.merge(df_db, on='Code', how='left')
            df_result['Task Name'] = df_result['Task Name'].fillna('')
            df_result['Grant ID'] = df_result['Grant ID'].fillna('')
            # Reordenar columnas: Code, Task Name, Grant ID, fechas...
            final_columns = ['Code', 'Task Name', 'Grant ID'] + date_columns
            df_result = df_result[final_columns]
        except Exception as e:
            print(f"Warning: Could not add Task Name/Grant ID: {e}")

    # Guardar resultado
    try:
        df_result.to_csv(output_path, index=False)
        print(f"Redistribution complete. Output saved to: {output_path}")
        print(f"Final result: {len(df_result)} rows")
        
        # Imprimir resumen
        virtual_total = df_virtual[date_columns].sum().sum()
        real_selected_total = df_real_selected[date_columns].sum().sum()
        real_not_selected_total = df_real_not_selected[date_columns].sum().sum() if len(df_real_not_selected) > 0 else 0
        excepted_total = df_excepted[date_columns].sum().sum() if len(df_excepted) > 0 else 0
        original_total = df_deltek[date_columns].sum().sum()
        final_total = df_result[date_columns].sum().sum()

        print(f"\n{'='*60}")
        print(f"REDISTRIBUTION SUMMARY")
        print(f"{'='*60}")
        print(f"ORIGINAL (before redistribution):")
        print(f"  - Virtual projects (to redistribute): {virtual_total:.2f}h")
        print(f"  - Real projects (selected): {real_selected_total:.2f}h")
        print(f"  - Real projects (NOT selected): {real_not_selected_total:.2f}h")
        print(f"  - Excepted projects: {excepted_total:.2f}h")
        print(f"  - TOTAL ORIGINAL: {original_total:.2f}h")
        print(f"\nFINAL (after redistribution):")
        print(f"  - TOTAL FINAL: {final_total:.2f}h")
        print(f"  - Difference: {final_total - original_total:.2f}h")
        print(f"  - Balance: {'✓ CONSERVED' if abs(final_total - original_total) < 0.1 else '✗ NOT CONSERVED'}")
        print(f"{'='*60}")
        
    except Exception as e:
        print(f"Error saving output file: {e}")


def show_prorate_comparison_window(original_file: str, prorated_file: str, database_path: str = None) -> bool:
    """
    Muestra ventana de comparación entre horas originales y con prorate por código de proyecto.

    Args:
        original_file (str): Ruta al archivo original 02-Timesheet.csv
        prorated_file (str): Ruta al archivo con prorate 03-Timesheet_Prorate.csv
        database_path (str): Ruta al archivo de base de datos con proyectos

    Returns:
        bool: True si el usuario acepta, False si cancela
    """
    try:
        # Cargar ambos archivos
        df_original = pd.read_csv(original_file)
        df_prorated = pd.read_csv(prorated_file)
        
        # Cargar información adicional de la base de datos de proyectos
        project_details = {}
        try:
            if database_path and os.path.exists(database_path):
                df_projects = pd.read_excel(database_path, sheet_name='N4W-Projects')

                # Crear diccionario con información del proyecto usando Code como clave
                for _, row in df_projects.iterrows():
                    task_name = row.get('Task Name', 'N/A')
                    if pd.isna(task_name) or task_name == '':
                        task_name = 'N/A'

                    project_details[row['Code']] = {
                        'Task_Name': task_name
                    }
            else:
                print(f"Warning: Database path not provided or file not found: {database_path}")
        except Exception as e:
            print(f"Warning: Could not load project database: {e}")
            project_details = {}

        # Obtener columnas de fechas
        date_columns_orig = get_date_columns(df_original)
        date_columns_pror = get_date_columns(df_prorated)

        # Agregar horas por Code (sumar todos los earnings)
        original_hours = df_original.groupby('Code')[date_columns_orig].sum().sum(axis=1)
        prorated_hours = df_prorated.groupby('Code')[date_columns_pror].sum().sum(axis=1)

        # Crear dataframe de comparación
        all_codes = set(original_hours.index) | set(prorated_hours.index)
        comparison_data = []

        for code in sorted(all_codes):
            orig_total = original_hours.get(code, 0.0)
            pror_total = prorated_hours.get(code, 0.0)
            
            # Obtener información adicional del proyecto
            details = project_details.get(code, {})
            
            comparison_data.append({
                'Code': code,
                'Task_Name': details.get('Task_Name', 'N/A'),
                'Without_Prorate': orig_total,
                'With_Prorate': pror_total,
                'Difference': pror_total - orig_total
            })

        # Crear ventana de comparación
        comparison_window = ctk.CTkToplevel()
        comparison_window.title("Prorate Hours Comparison")
        comparison_window.geometry("1000x600")
        comparison_window.configure(fg_color=COLORS['bg_primary'])
        comparison_window.transient()
        comparison_window.grab_set()  # Hacer ventana modal

        # Centrar la ventana
        comparison_window.update_idletasks()
        x = (comparison_window.winfo_screenwidth() // 2) - (1000 // 2)
        y = (comparison_window.winfo_screenheight() // 2) - (600 // 2)
        comparison_window.geometry(f"1000x600+{x}+{y}")

        # Variable para almacenar la elección del usuario
        user_choice = {'accepted': False}

        def on_accept():
            user_choice['accepted'] = True
            comparison_window.destroy()

        def on_cancel():
            user_choice['accepted'] = False
            comparison_window.destroy()

        # Encabezado
        header_frame = ctk.CTkFrame(comparison_window, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(20, 10))

        title_label = ctk.CTkLabel(
            header_frame,
            text="Hours Distribution Comparison",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=COLORS['text_primary']
        )
        title_label.pack()

        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Review hour redistribution by project. Total hours are conserved - only distribution changes.",
            font=ctk.CTkFont(size=14),
            text_color=COLORS['text_secondary']
        )
        subtitle_label.pack(pady=(5, 0))

        # Marco de tabla con vista desplazable
        table_frame = ctk.CTkScrollableFrame(
            comparison_window,
            fg_color=COLORS['bg_secondary'],
            corner_radius=8,
            border_width=1,
            border_color=COLORS['border']
        )
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Encabezados de tabla con anchos fijos
        headers_frame = ctk.CTkFrame(table_frame, fg_color=COLORS['bg_tertiary'])
        headers_frame.pack(fill="x", pady=(0, 5))

        # Definir anchos fijos para columnas
        col_widths = [80, 450, 80, 80, 80]

        ctk.CTkLabel(headers_frame, text="Code", font=ctk.CTkFont(weight="bold"), width=col_widths[0]).grid(row=0, column=0, padx=2, pady=8, sticky="w")
        ctk.CTkLabel(headers_frame, text="Task Name", font=ctk.CTkFont(weight="bold"), width=col_widths[1]).grid(row=0, column=1, padx=2, pady=8, sticky="w")
        ctk.CTkLabel(headers_frame, text="Original", font=ctk.CTkFont(weight="bold"), width=col_widths[2]).grid(row=0, column=2, padx=2, pady=8, sticky="w")
        ctk.CTkLabel(headers_frame, text="Prorated", font=ctk.CTkFont(weight="bold"), width=col_widths[3]).grid(row=0, column=3, padx=2, pady=8, sticky="w")
        ctk.CTkLabel(headers_frame, text="Diff", font=ctk.CTkFont(weight="bold"), width=col_widths[4]).grid(row=0, column=4, padx=2, pady=8, sticky="w")

        # Filas de tabla
        for i, row_data in enumerate(comparison_data):
            row_color = COLORS['bg_primary'] if i % 2 == 0 else COLORS['bg_secondary']

            row_frame = ctk.CTkFrame(table_frame, fg_color=row_color, corner_radius=4)
            row_frame.pack(fill="x", pady=1)

            # Resaltar proyectos virtuales (los que van a 0)
            text_color = COLORS['warning'] if row_data['With_Prorate'] == 0 and row_data['Without_Prorate'] > 0 else \
            COLORS['text_primary']

            # Usar los mismos anchos fijos que los encabezados
            ctk.CTkLabel(row_frame, text=row_data['Code'], text_color=text_color, width=col_widths[0]).grid(row=0, column=0, padx=2, pady=6, sticky="w")
            ctk.CTkLabel(row_frame, text=row_data['Task_Name'], text_color=text_color, width=col_widths[1]).grid(row=0, column=1, padx=2, pady=6, sticky="w")
            ctk.CTkLabel(row_frame, text=f"{row_data['Without_Prorate']:.1f}", text_color=text_color, width=col_widths[2]).grid(row=0, column=2, padx=2, pady=6, sticky="w")
            ctk.CTkLabel(row_frame, text=f"{row_data['With_Prorate']:.1f}", text_color=text_color, width=col_widths[3]).grid(row=0, column=3, padx=2, pady=6, sticky="w")

            # Colorear la diferencia
            diff = row_data['Difference']
            diff_color = COLORS['success'] if diff > 0 else (
                COLORS['warning'] if diff < 0 else COLORS['text_secondary'])
            diff_text = f"+{diff:.1f}" if diff > 0 else f"{diff:.1f}"
            ctk.CTkLabel(row_frame, text=diff_text, text_color=diff_color, width=col_widths[4]).grid(row=0, column=4, padx=2, pady=6, sticky="w")

        # Marco de resumen
        summary_frame = ctk.CTkFrame(comparison_window, fg_color=COLORS['bg_tertiary'], corner_radius=8)
        summary_frame.pack(fill="x", padx=20, pady=10)

        original_total = sum(data['Without_Prorate'] for data in comparison_data)
        prorated_total = sum(data['With_Prorate'] for data in comparison_data)

        hours_match = abs(original_total - prorated_total) < 0.01
        match_symbol = '✓' if hours_match else '✗'
        match_color = COLORS['success'] if hours_match else COLORS['warning']

        summary_text = f"TOTAL HOURS - Original: {original_total:.2f}h | After Redistribution: {prorated_total:.2f}h | Balance: {match_symbol}"
        ctk.CTkLabel(
            summary_frame,
            text=summary_text,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=match_color
        ).pack(pady=10)

        # Leyenda
        legend_frame = ctk.CTkFrame(comparison_window, fg_color="transparent")
        legend_frame.pack(fill="x", padx=20)

        ctk.CTkLabel(
            legend_frame,
            text="🟡 Pro-rated projects (redistributed to other projects)",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['warning']
        ).pack(side="left")

        # Marco de botones
        button_frame = ctk.CTkFrame(comparison_window, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=(10, 20))

        cancel_button = ctk.CTkButton(
            button_frame,
            text="Cancel Process",
            command=on_cancel,
            width=120,
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#DC2626",
            hover_color="#B91C1C"
        )
        cancel_button.pack(side="left")

        accept_button = ctk.CTkButton(
            button_frame,
            text="Create File",
            command=on_accept,
            width=150,
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLORS['success'],
            hover_color='#0D6A0D'
        )
        accept_button.pack(side="right")

        # Esperar elección del usuario
        comparison_window.wait_window()

        return user_choice['accepted']

    except Exception as e:
        print(f"Error creating comparison window: {e}")
        # Si hay error, por defecto continuar
        return True

# =============================================================================
# FUNCIÓN PARA MOSTRAR PROYECTOS ELIMINADOS
# =============================================================================
def show_removed_projects_window(proyectos_a_eliminar):
    """
    Muestra una ventana con scroll que lista los proyectos eliminados.

    Args:
        proyectos_a_eliminar (list): Lista de diccionarios con info de proyectos eliminados
    """
    # Crear ventana
    window = ctk.CTkToplevel()
    window.title("Projects Removed from Database")
    window.geometry("700x500")
    window.configure(fg_color=COLORS['bg_primary'])

    # Centrar la ventana
    window.update_idletasks()
    x = (window.winfo_screenwidth() // 2) - (700 // 2)
    y = (window.winfo_screenheight() // 2) - (500 // 2)
    window.geometry(f"700x500+{x}+{y}")

    # Hacer ventana modal
    window.transient()
    window.grab_set()

    # Encabezado
    header_frame = ctk.CTkFrame(window, fg_color="transparent")
    header_frame.pack(fill="x", padx=20, pady=(20, 10))

    title_label = ctk.CTkLabel(
        header_frame,
        text="Projects Removed from Database",
        font=ctk.CTkFont(size=20, weight="bold"),
        text_color=COLORS['text_primary']
    )
    title_label.pack()

    subtitle_label = ctk.CTkLabel(
        header_frame,
        text=f"Total removed: {len(proyectos_a_eliminar)} project(s)",
        font=ctk.CTkFont(size=13),
        text_color=COLORS['text_secondary']
    )
    subtitle_label.pack(pady=(5, 0))

    # Frame con scroll para la lista de proyectos
    scrollable_frame = ctk.CTkScrollableFrame(
        window,
        fg_color=COLORS['bg_secondary'],
        corner_radius=8,
        border_width=1,
        border_color=COLORS['border']
    )
    scrollable_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))

    # Lista de proyectos eliminados
    for i, proyecto in enumerate(proyectos_a_eliminar):
        # Frame para cada proyecto
        project_frame = ctk.CTkFrame(
            scrollable_frame,
            fg_color=COLORS['bg_tertiary'],
            corner_radius=6
        )
        project_frame.pack(fill="x", padx=5, pady=5)

        # Código del proyecto
        code_label = ctk.CTkLabel(
            project_frame,
            text=f"• {proyecto['code']}",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS['warning'],
            anchor="w"
        )
        code_label.pack(fill="x", padx=15, pady=(10, 2))

        # Descripción
        desc_label = ctk.CTkLabel(
            project_frame,
            text=proyecto['description'],
            font=ctk.CTkFont(size=12),
            text_color=COLORS['text_primary'],
            anchor="w"
        )
        desc_label.pack(fill="x", padx=30, pady=2)

        # Razón
        reason_label = ctk.CTkLabel(
            project_frame,
            text=f"Reason: {proyecto['reason']}",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['text_secondary'],
            anchor="w"
        )
        reason_label.pack(fill="x", padx=30, pady=(2, 10))

    # Botón OK
    ok_button = ctk.CTkButton(
        window,
        text="OK",
        command=window.destroy,
        width=100,
        height=36,
        font=ctk.CTkFont(size=13, weight="bold"),
        fg_color=COLORS['accent'],
        hover_color=COLORS['accent_hover']
    )
    ok_button.pack(pady=(0, 20))

    # Esperar a que se cierre la ventana
    window.wait_window()

# =============================================================================
# FUNCIONES DE ACTUALIZACIÓN DE BASE DE DATOS
# =============================================================================
def refresh_excel_formulas(file_path):
    file_path = os.path.abspath(file_path)

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False

    try:
        wb = xl.Workbooks.Open(file_path)
        xl.CalculateUntilAsyncQueriesDone()
        wb.Save()
        wb.Close()
    finally:
        xl.Quit()

def Update_DataBase_With_BoxFile(archivo_base, archivo_fuente):
    """
    Actualiza la base de datos local con información del archivo de N4W de Box.
    Solo se ejecuta una vez por sesión para evitar actualizaciones duplicadas.

    Args:
        archivo_base (str): Ruta del archivo de base de datos local
        archivo_fuente (str): Ruta del archivo fuente descargado de Box
    """
    print("=" * 70)
    print("UPDATING DATABASE FOR THE FIRST TIME IN THIS SESSION")
    print("=" * 70)

    # Convertir a ruta absoluta al inicio
    archivo_base = os.path.abspath(archivo_base)
    password = "TimeSheet_N4W"

    # Función para verificar si un valor está "vacío"
    def esta_vacio(valor):
        if pd.isna(valor):
            return True
        if valor is None:
            return True
        if isinstance(valor, str) and len(valor.strip()) == 0:
            return True
        return False

    # ============================================================================
    # PASO 0: LEER TODOS LOS DATOS CON PANDAS (ANTES DE OPERACIONES COM)
    # ============================================================================
    print("\n[PASO 0] Reading data with pandas...")
    df_base = pd.read_excel(archivo_base, sheet_name='N4W-Projects')
    df_fuente = pd.read_excel(archivo_fuente)

    print(f"Rows in base: {len(df_base)}")
    print(f"Rows in source: {len(df_fuente)}")
    print(f"Columns in df_base: {df_base.columns.tolist()}")

    # ============================================================================
    # PASO 1: PROCESAR ACTUALIZACIONES DE PROYECTOS (EN MEMORIA)
    # ============================================================================
    print("\n[PASO 1] Processing project updates...")

    # Función para verificar si código es especial (empieza con "XX")
    def es_codigo_especial(code):
        return str(code).strip().upper().startswith('XX')

    # Obtener solo los Code que NO están vacíos y NO son especiales (XX)
    CodeN4W_ids_validos1 = set()
    for idx in df_base.index:
        CodeN4W_id = df_base.loc[idx, 'Code']
        if not esta_vacio(CodeN4W_id) and not es_codigo_especial(CodeN4W_id):
            CodeN4W_ids_validos1.add(CodeN4W_id)

    task_names = set(df_fuente['Task_Name'].dropna())

    # Verificar que todos los códigos válidos de la base existan en el fuente (excluir XX y -1)
    CodeN4W_ids_validos = {x for x in CodeN4W_ids_validos1
                          if not (isinstance(x, (int, float)) and not isinstance(x, bool) and x == -1)
                          and not es_codigo_especial(x)}

    faltantes = CodeN4W_ids_validos - task_names

    if faltantes:
        raise ValueError(f"ERROR: The following Code were not found in the source file: {faltantes}")

    # Crear diccionario para mapear los datos
    df_fuente_indexed = df_fuente.set_index('Task_Name')

    # Actualizar las columnas fila por fila (ignorar códigos XX)
    for idx in df_base.index:
        CodeN4W_id = df_base.loc[idx, 'Code']

        # Ignorar códigos especiales (XX)
        if es_codigo_especial(CodeN4W_id):
            print(f"  → Ignoring special code: {CodeN4W_id}")
            continue

        if not esta_vacio(CodeN4W_id) and CodeN4W_id in df_fuente_indexed.index:
            # Actualizar solo Task Name y Grant ID
            df_base.loc[idx, 'Description'] = df_fuente_indexed.loc[CodeN4W_id, 'Task_Name_Description']
            df_base.loc[idx, 'Task Name'] = df_fuente_indexed.loc[CodeN4W_id, 'WD_TaskName']
            df_base.loc[idx, 'Grant ID'] = df_fuente_indexed.loc[CodeN4W_id, 'WD_GrantID']

            # Actualizar Category (concatenación de Code | Description)
            df_base.loc[idx, 'Category'] = f"{CodeN4W_id} | {df_base.loc[idx, 'Description']}"

        elif esta_vacio(CodeN4W_id):
            # Si Code está vacío, poner en "0"
            df_base.loc[idx, 'Description'] = "0"
            df_base.loc[idx, 'Task Name'] = "0"
            df_base.loc[idx, 'Grant ID'] = "0"
            df_base.loc[idx, 'Category'] = "0"

    # ============================================================================
    # PASO 2: IDENTIFICAR Y ELIMINAR PROYECTOS CERRADOS (EN MEMORIA)
    # ============================================================================
    print("\n[PASO 2] Identifying closed projects...")
    proyectos_a_eliminar = []
    indices_a_eliminar = []

    # Iterar sobre la BASE DE DATOS (no sobre df_fuente)
    for idx in df_base.index:
        code_actual = df_base.loc[idx, 'Code']

        # Ignorar códigos especiales (XX) - no eliminar ni procesar
        if es_codigo_especial(code_actual):
            continue

        if code_actual == -1:
            # Condición 3: Valor de -1 en Code
            indices_a_eliminar.append(idx)

        # Solo procesar códigos que NO estén vacíos
        if not esta_vacio(code_actual):
            # Verificar si este código existe en df_fuente
            if code_actual in df_fuente_indexed.index:
                # Obtener las fechas de este proyecto en df_fuente
                date_opening = df_fuente_indexed.loc[
                    code_actual, 'Date_Opened'] if 'Date_Opened' in df_fuente_indexed.columns else None
                date_closing = df_fuente_indexed.loc[
                    code_actual, 'Date_Closed'] if 'Date_Closed' in df_fuente_indexed.columns else None

                # Condición 1: Date_Opened vacío
                sin_apertura = esta_vacio(date_opening)

                # Condición 2: Date_Closed con valor
                con_cierre = not esta_vacio(date_closing)

                # Si cumple alguna de las dos condiciones, marcar para eliminar
                if sin_apertura or con_cierre:
                    razon = []
                    if sin_apertura:
                        razon.append("No Date_Opened")
                    if con_cierre:
                        razon.append("Has Date_Closed")

                    # Obtener descripción del proyecto
                    descripcion = df_base.loc[idx, 'Description']

                    proyectos_a_eliminar.append({
                        'code': code_actual,
                        'description': descripcion,
                        'reason': ' | '.join(razon)
                    })

                    # Marcar índice para eliminar
                    indices_a_eliminar.append(idx)

                    print(f"  → Marked for deletion: {code_actual} - {descripcion} ({' | '.join(razon)})")

    print(f"\nTotal projects to remove: {len(proyectos_a_eliminar)}")
    print(f"Rows before deletion: {len(df_base)}")

    # Eliminar las filas del DataFrame
    if indices_a_eliminar:
        df_base = df_base.drop(indices_a_eliminar)
        df_base = df_base.reset_index(drop=True)

        print(f"Rows after deletion: {len(df_base)}")
        print(f"Removed {len(indices_a_eliminar)} projects from database")

        # Mostrar ventana personalizada con scroll para los proyectos eliminados
        show_removed_projects_window(proyectos_a_eliminar)
    else:
        print("No projects to remove")

    # ============================================================================
    # PASO 3: ESCRIBIR TODOS LOS CAMBIOS A EXCEL (UNA SOLA INSTANCIA COM)
    # ============================================================================
    print("\n[PASO 3] Writing changes to Excel with single COM instance...")
    xl = None
    wb = None

    try:
        # Crear UNA SOLA instancia Excel COM
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False

        # Abrir el workbook
        try:
            wb = xl.Workbooks.Open(archivo_base, Password=password)
        except:
            wb = xl.Workbooks.Open(archivo_base)

        ws = wb.Worksheets('N4W-Projects')

        # Verificar si la hoja está protegida y desprotegerla
        sheet_was_protected = ws.ProtectContents
        if sheet_was_protected:
            ws.Unprotect(password)
            print("Sheet unprotected successfully")

        # Escribir los datos actualizados (todas las columnas para mantener sincronización)
        for idx, row in df_base.iterrows():
            fila_excel = idx + 2  # +2 porque índice empieza en 0 y hay encabezado
            ws.Cells(fila_excel, 1).Value = row['Code']  # Columna A
            ws.Cells(fila_excel, 2).Value = row['Description']  # Columna B
            ws.Cells(fila_excel, 3).Value = row['Task Name']  # Columna C
            ws.Cells(fila_excel, 4).Value = row['Grant ID']  # Columna D
            ws.Cells(fila_excel, 5).Value = row['Category']  # Columna E

            # Escribir Include si existe en el DataFrame
            if 'Include' in df_base.columns:
                ws.Cells(fila_excel, 6).Value = row['Include']  # Columna F

        # Si se eliminaron filas del DataFrame, borrar las filas sobrantes del Excel
        ultima_fila_excel = ws.UsedRange.Rows.Count
        filas_en_dataframe = len(df_base) + 1  # +1 por el encabezado

        if ultima_fila_excel > filas_en_dataframe:
            print(f"Deleting {ultima_fila_excel - filas_en_dataframe} extra rows from Excel")
            for fila in range(ultima_fila_excel, filas_en_dataframe, -1):
                ws.Rows(fila).Delete()
            print(f"Successfully deleted extra rows from Excel")

        # Volver a proteger la hoja si estaba protegida
        if sheet_was_protected:
            ws.Protect(password)
            print("Sheet protected again successfully")

        # Guardar cambios
        wb.Save()
        print("Changes saved successfully")

        # ========================================================================
        # PASO 4: REFRESH FORMULAS (MISMA INSTANCIA COM)
        # ========================================================================
        print("\n[PASO 4] Refreshing Excel formulas...")
        xl.CalculateUntilAsyncQueriesDone()
        wb.Save()
        print("Formulas refreshed successfully")

        # Cerrar workbook
        wb.Close()
        wb = None
        print("Workbook closed successfully")

    except Exception as e:
        print(f"Error during Excel COM operation: {e}")
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except:
            pass
        raise

    finally:
        # GARANTIZAR liberación de Excel COM
        try:
            if xl is not None:
                xl.Quit()
                print("Excel COM instance released")
        except Exception as e:
            print(f"Warning: Error releasing Excel COM: {e}")

    print("\n" + "=" * 70)
    print("DATABASE UPDATE COMPLETED SUCCESSFULLY")
    print("=" * 70)

    # Mostrar messagebox de actualización exitosa
    messagebox.showinfo("Database Updated",
                        "Database updated successfully!\n\nAll project information has been synchronized.")



def Download_DataBase_N4W_Box(url_box, salida):
    """
    Descarga la base de datos de N4W desde Box.
    
    Args:
        url_box (str): URL del archivo en Box
        salida (str): Ruta donde guardar el archivo descargado
    """
    # Convertir URL de preview a URL de descarga directa
    url_descarga = url_box.replace('/s/', '/shared/static/')

    # Descargar el archivo
    response = requests.get(url_descarga)
    response.raise_for_status()  # Verificar que la descarga fue exitosa

    # Guardar el archivo
    with open(salida, 'wb') as archivo:
        archivo.write(response.content)

    print(f"File downloaded successfully to: {salida}")


def readDataBase(filepath):
    """
    Lee y combina datos de múltiples hojas de Excel.
    
    Args:
        filepath (str): Ruta al archivo Excel
        
    Returns:
        pd.DataFrame: Datos combinados de todas las hojas
    """
    #df1 = pd.read_excel(filepath, sheet_name='TNC-Employee')
    df2 = pd.read_excel(filepath, sheet_name='N4W-Projects')
    #df3 = pd.read_excel(filepath, sheet_name='TNC-Projects')
    # return pd.concat([df1, df2, df3], ignore_index=True)
    return df2


def Lookup_UserName_Outlook(email: str) -> Optional[Dict[str, str]]:
    """
    Busca el nombre de la persona asociada a un correo en Outlook.
    - Primero intenta el directorio (Global Address List) si hay cuenta Exchange/365.
    - Si no, busca en Contactos locales (Email1/Email2/Email3).
    - Devuelve dict con name, email y metadatos cuando Outlook los expone.

    Args:
        email (str): Dirección de correo electrónico a buscar

    Returns:
        Optional[Dict[str, str]]: Diccionario con información del usuario o None
    """
    outlook = None
    session = None

    # Resultado base
    result = {"email": email, "name": None}

    try:
        # Inicia Outlook (o se conecta a una instancia existente)
        outlook = win32com.client.Dispatch("Outlook.Application")
        session = outlook.Session  # MAPI Namespace

        # --- 1) Resolver en directorio (Exchange/365) ---
        # CreateRecipient intenta resolver en GAL/Directorio si existe
        recipient = session.CreateRecipient(email)
        recipient.Resolve()
        if recipient.Resolved:
            ae = recipient.AddressEntry
            # Nombre "display" genérico
            result["name"] = ae.Name

            # Si es usuario Exchange, podemos sacar datos más ricos
            try:
                ex_user = ae.GetExchangeUser()
            except Exception:
                ex_user = None

            if ex_user:
                # ex_user.PrimarySmtpAddress suele ser el correo "real"
                result.update({
                    "email": ex_user.PrimarySmtpAddress or email,
                    "name": ex_user.Name or ae.Name or None,
                })
                # Extra opcional si está disponible
                try:
                    if ex_user.JobTitle:
                        result["job_title"] = ex_user.JobTitle
                except Exception:
                    pass
                try:
                    if ex_user.CompanyName:
                        result["company"] = ex_user.CompanyName
                except Exception:
                    pass
                return result

            # Si no es ExchangeUser (p.ej. contacto de Internet), vale el display name
            if result["name"]:
                return result

        # --- 2) Buscar en Contactos locales ---
        try:
            contacts = session.GetDefaultFolder(constants.olFolderContacts)  # 10
            items = contacts.Items
            # Revisamos hasta 3 campos de email que Outlook maneja en Contactos
            for field in ("Email1Address", "Email2Address", "Email3Address"):
                # Items.Find usa la sintaxis de restricción de Outlook
                found = items.Find(f"[{field}] = '{email}'")
                if found:
                    result["name"] = getattr(found, "FullName", None) or getattr(found, "CompanyName", None)
                    # Si Outlook almacenó el email con normalización distinta, respétalo
                    try:
                        normalized = getattr(found, field, None)
                        if normalized:
                            result["email"] = normalized
                    except Exception:
                        pass
                    return result
        except Exception:
            # Si no hay carpeta de contactos o no se puede acceder, continuamos
            pass

        # --- 3) Último intento: "resolver" sólo para obtener display name genérico ---
        if not recipient.Resolved:
            # A veces Resolve falla con el correo; probamos con Recipient de nuevo
            recipient = session.CreateRecipient(email)
            recipient.Resolve()
        if recipient and recipient.Resolved:
            result["name"] = recipient.Name or result["name"]

        # Si llegamos aquí, devolvemos lo que tengamos (quizá sólo el email)
        return result if (result.get("name") or result.get("email")) else None

    except Exception as e:
        # Algo muy raro (p.ej. Outlook no configurado)
        raise RuntimeError(f"Unable to access Outlook: {e}")

    finally:
        # Liberar objetos Outlook COM
        try:
            if session is not None:
                session = None
            if outlook is not None:
                outlook = None
        except Exception as e:
            print(f"Warning: Error releasing Outlook COM: {e}")


# =============================================================================
# FUNCIONES DE ONEDRIVE
# =============================================================================

def _env_onedrive_candidates() -> List[Path]:
    """Obtiene candidatos desde variables de entorno estándar de OneDrive."""
    candidates = []
    for var in ("OneDriveCommercial", "OneDriveConsumer", "OneDrive"):
        p = os.environ.get(var)
        if p:
            pp = Path(p).expanduser().resolve()
            if pp.exists():
                candidates.append(pp)
    # Deduplicar preservando orden
    uniq = []
    seen = set()
    for c in candidates:
        if c not in seen:
            uniq.append(c)
            seen.add(c)
    return uniq


def _registry_onedrive_accounts() -> List[Dict[str, str]]:
    """
    Lee el registro para descubrir cuentas de OneDrive.
    Devuelve lista de dicts: {"display_name": str, "user_folder": str, "kind": "Personal"/"BusinessN"}
    """
    results: List[Dict[str, str]] = []
    if winreg is None:
        return results

    base_path = r"Software\Microsoft\OneDrive\Accounts"
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, base_path) as accounts:
            i = 0
            while True:
                try:
                    subname = winreg.EnumKey(accounts, i)  # p.ej., "Personal", "Business1", "Business2"
                    i += 1
                except OSError:
                    break

                # Abrir subclave
                try:
                    with winreg.OpenKey(accounts, subname) as subkey:
                        user_folder = _reg_get_str(subkey, "UserFolder")
                        display_name = _reg_get_str(subkey, "DisplayName") or _pretty_label_from_path(user_folder)
                        if user_folder:
                            results.append({
                                "display_name": display_name,
                                "user_folder": user_folder,
                                "kind": subname
                            })
                except OSError:
                    continue
    except OSError:
        pass
    return results


def _reg_get_str(key, value_name: str) -> Optional[str]:
    """Obtiene valor string del registro."""
    try:
        val, typ = winreg.QueryValueEx(key, value_name)
        if isinstance(val, str):
            return val
    except OSError:
        return None
    return None


def _pretty_label_from_path(p: Optional[str]) -> Optional[str]:
    """Genera etiqueta legible desde ruta."""
    if not p:
        return None
    base = Path(p).name
    # Normalmente algo como "OneDrive - MiEmpresa" o "OneDrive"
    return base


def get_onedrive_accounts() -> List[Dict[str, str]]:
    """
    Devuelve lista de cuentas OneDrive encontradas:
    [
      {"label": "OneDrive - MiEmpresa", "root": "C:\\Users\\yo\\OneDrive - MiEmpresa"},
      {"label": "OneDrive", "root": "C:\\Users\\yo\\OneDrive"},
      ...
    ]
    Combina registro y variables de entorno, deduplicando.
    """
    accounts: List[Dict[str, str]] = []

    # 1) Registro (más confiable)
    for acc in _registry_onedrive_accounts():
        root = Path(acc["user_folder"]).expanduser().resolve()
        if root.exists():
            label = _pretty_label_from_path(str(root))
            accounts.append({"label": label, "root": str(root)})

    # 2) Variables de entorno (por si faltó algo)
    for p in _env_onedrive_candidates():
        label = _pretty_label_from_path(str(p))
        entry = {"label": label, "root": str(p)}
        if entry not in accounts:
            accounts.append(entry)

    # Dedup por root
    seen = set()
    uniq = []
    for a in accounts:
        if a["root"] not in seen:
            uniq.append(a)
            seen.add(a["root"])
    return uniq


def _split_on_first(path_str: str) -> List[str]:
    """Divide ruta en partes."""
    parts = re.split(r"[\\/]+", path_str.strip().strip("\\/"))
    return [p for p in parts if p]


def resolve_onedrive_target(target_path_in_onedrive: str,
                            account_hint: Optional[str] = None) -> Path:
    """
    Decide la raíz de OneDrive y construye la ruta destino absoluta.
    """
    parts = _split_on_first(target_path_in_onedrive)
    if not parts:
        raise ValueError("Destination route empty.")

    accounts = get_onedrive_accounts()
    if not accounts:
        raise RuntimeError("No OneDrive folder was detected in this Windows profile.")

    # ¿El primer segmento coincide exactamente con el label de alguna cuenta?
    first = parts[0]
    by_label = {a["label"]: a for a in accounts if a["label"]}
    if first in by_label:
        root = Path(by_label[first]["root"])
        rel_parts = parts[1:]
        return root.joinpath(*rel_parts)

    # Si no coincide, elegir cuenta
    chosen = None
    if len(accounts) == 1:
        chosen = accounts[0]
    else:
        # Intentar por pista
        if account_hint:
            hint = account_hint.lower()
            # match por label o por nombre de carpeta base
            candidates = [a for a in accounts
                          if (a["label"] and hint in a["label"].lower()) or
                             (hint in Path(a["root"]).name.lower())]
            if len(candidates) == 1:
                chosen = candidates[0]
            elif len(candidates) > 1:
                # si hay varias, preferir la que contenga más "match" (heurística simple)
                candidates.sort(key=lambda a: (a["label"] or "").lower().find(hint))
                chosen = candidates[0]
        # Si sigue sin elegirse, intentar heurística por patrón "OneDrive - <algo>"
        if not chosen:
            enterprise = [a for a in accounts if " - " in (a["label"] or "")]
            chosen = enterprise[0] if enterprise else accounts[0]

    Tmp = os.path.split(chosen["root"]) #chosen["root"].replace("OneDrive - ", "")
    chosen["root"] = os.path.join(Tmp[0],"The Nature Conservancy")
    root = Path(chosen["root"])
    return root.joinpath(*parts)


def put_file_in_onedrive(src_path: str,
                         target_path_in_onedrive: str,
                         account_hint: Optional[str] = None,
                         move: bool = False,
                         overwrite: bool = False) -> Path:
    """
    Copia o mueve un archivo a OneDrive.
    """
    src = Path(src_path).expanduser().resolve()
    if not src.exists():
        raise FileNotFoundError(f"The source file does not exist: {src}")
    if not src.is_file():
        raise IsADirectoryError(f"The origin is not a file: {src}")

    dst = resolve_onedrive_target(target_path_in_onedrive, account_hint=account_hint)
    dst_parent = dst.parent
    dst_parent.mkdir(parents=True, exist_ok=True)

    if dst.exists():
        if overwrite:
            # Intento de borrado seguro (si es archivo)
            if dst.is_file():
                try:
                    os.remove(dst)
                except PermissionError:
                    # Si está bloqueado (p.ej. sincronizando), renombramos el viejo antes
                    backup = dst.with_suffix(dst.suffix + f".bak.{uuid.uuid4().hex[:8]}")
                    dst.replace(backup)
            else:
                raise IsADirectoryError(f"A folder already exists at the destination: {dst}")
        else:
            raise FileExistsError(f"The destination file already exists: {dst}")

    if move:
        # shutil.move maneja discos distintos.
        shutil.move(str(src), str(dst))
    else:
        # Remplazar para que sea solo la ruta compartida por Sunil
        dst = str(dst).replace('OneDrive - ', '')
        dst = dst.replace('OneDrive', '')
        shutil.copy2(str(src), str(dst))  # conserva metadata básica

    return dst


# =============================================================================
# CONFIGURACIÓN DE ARCHIVO EN FORMATO POWERAPP - TIMESHEET N4W
# =============================================================================
def cargar_base_datos_tareas(archivo_base_datos):
    """
    Carga el archivo de base de datos de tareas y crea un diccionario de búsqueda.

    Parámetros:
    - archivo_base_datos: ruta del archivo "N4W Task Details.xlsx"

    Retorna:
    - diccionario con Task_Name como clave y Task_Name_Description como valor
    """
    try:
        # Leer el archivo Excel de base de datos
        df_base = pd.read_excel(archivo_base_datos, sheet_name='Task_Details')

        # Crear diccionario de búsqueda: Task_Name -> Task_Name_Description
        diccionario_tareas = dict(zip(df_base['Task_Name'], df_base['Timesheet Code']))

        print(f"Database loaded: {len(diccionario_tareas)} tasks found")
        return diccionario_tareas

    except Exception as e:
        print(f"Error loading database: {e}")
        return {}


def CreateExcel_N4WFormat(archivo_csv, email_empleado, nombre_empleado, ruta_guardado, archivo_base_datos=None,
                          NameTableSheet='new_n4wtimeentriessubmissionses'):
    """
    Convierte datos de CSV de Deltek a formato Excel de timesheet con tabla de Excel.

    Parámetros:
    - archivo_csv: ruta del archivo CSV de entrada
    - email_empleado: email del empleado
    - nombre_empleado: nombre del empleado
    - ruta_guardado: ruta completa donde guardar el archivo Excel
    - archivo_base_datos: ruta del archivo "N4W Task Details.xlsx" (opcional)
    """

    # Cargar base de datos de tareas si se proporciona
    diccionario_tareas = {}
    if archivo_base_datos and os.path.exists(archivo_base_datos):
        diccionario_tareas = cargar_base_datos_tareas(archivo_base_datos)

    # Leer el CSV
    df = pd.read_csv(archivo_csv)

    # Reemplazar códigos XX por OF0104
    df.loc[df['Code'].astype(str).str.upper().str.startswith('XX'), 'Code'] = 'OF0104'

    # Obtener las columnas de fechas (todas las que tienen formato de fecha)
    columnas_fecha = [col for col in get_date_columns(df) if '00:00:00' in col]

    # Convertir columnas de fecha a datetime
    fechas = []
    for col in columnas_fecha:
        fecha_str = col.replace(' 00:00:00', '')
        fechas.append(datetime.strptime(fecha_str, '%Y-%m-%d'))

    # Filtrar filas: eliminar las que tengan Code que inicien con "TNC"
    df_filtrado = df[~df['Code'].str.startswith('TNC', na=False)]

    # Agrupar por proyecto y sumar las horas
    columnas_agrupacion = ['Code']
    df_agrupado = df_filtrado.groupby(columnas_agrupacion, as_index=False)[columnas_fecha].sum()

    # Crear lista para almacenar las filas del Excel final
    filas_excel = []

    # Procesar cada proyecto
    for _, fila in df_agrupado.iterrows():
        CodeN4W_id = fila['Code']  # Usar Code

        # Buscar la descripción en la base de datos
        if diccionario_tareas and CodeN4W_id in diccionario_tareas:
            codigo_proyecto = diccionario_tareas[CodeN4W_id]
        else:
            # Si no se encuentra, usar el Code original
            codigo_proyecto = CodeN4W_id
            if diccionario_tareas:  # Solo mostrar advertencia si se cargó la base de datos
                print(f"Warning: Description not found for {CodeN4W_id}")

        # Agrupar datos por semanas
        datos_por_semana = {}

        for i, fecha in enumerate(fechas):
            col_fecha = columnas_fecha[i]
            horas = fila[col_fecha] if pd.notna(fila[col_fecha]) else 0

            # Encontrar el domingo de esa semana (inicio de semana)
            dias_desde_lunes = fecha.weekday()
            inicio_semana = fecha - timedelta(days=dias_desde_lunes)

            # dias_desde_domingo = (fecha.weekday() + 1) % 7  # Domingo = 0
            # inicio_semana = fecha - timedelta(days=dias_desde_domingo)

            if inicio_semana not in datos_por_semana:
                datos_por_semana[inicio_semana] = {
                    'sun': 0, 'mon': 0, 'tue': 0, 'wed': 0,
                    'thu': 0, 'fri': 0, 'sat': 0
                }

            # Asignar horas al día correspondiente
            dias_semana = ['sun', 'mon', 'tue', 'wed', 'thu', 'fri', 'sat']
            dia_semana = dias_semana[(fecha.weekday() + 1) % 7]
            datos_por_semana[inicio_semana][dia_semana] += horas

        # Crear filas para cada semana que tenga horas
        for inicio_semana, horas_semana in datos_por_semana.items():
            total_horas = sum(horas_semana.values())

            if total_horas > 0:  # Solo incluir semanas con horas trabajadas
                fin_semana = inicio_semana + timedelta(days=6)

                # Formatear título de la semana
                titulo_semana = f"{inicio_semana.strftime('%d-%B-%Y')} to {fin_semana.strftime('%d-%B-%Y')}"

                # Formatear fecha de inicio como objeto datetime (no como string)
                fecha_excel = inicio_semana

                fila_excel = {
                    'new_title': titulo_semana,
                    'new_employeeemail': email_empleado,
                    'new_employeename': nombre_empleado,
                    'new_projectcode': codigo_proyecto,
                    'new_monhours': horas_semana['mon'],
                    'new_tuehours': horas_semana['tue'],
                    'new_wedhours': horas_semana['wed'],
                    'new_thurshours': horas_semana['thu'],
                    'new_frihours': horas_semana['fri'],
                    'new_sathours': horas_semana['sat'],
                    'new_sunhours': horas_semana['sun'],
                    'new_totalhours': total_horas,
                    'crd63_timesheetinitiated': True,
                    'new_timesheetstatus': 'Submitted',
                    'crd63_weekstartdate': fecha_excel,
                    'new_comments': 'Submitted'
                }

                filas_excel.append(fila_excel)

    # Crear DataFrame final
    df_final = pd.DataFrame(filas_excel)

    # Ordenar por fecha de inicio de semana
    df_final = df_final.sort_values('crd63_weekstartdate')

    # Crear workbook de Excel usando openpyxl directamente para tener control total
    wb = Workbook()
    ws = wb.active
    ws.title = NameTableSheet

    # Escribir los encabezados
    encabezados = [
        'new_title', 'new_employeeemail', 'new_employeename', 'new_projectcode',
        'new_monhours', 'new_tuehours', 'new_wedhours', 'new_thurshours',
        'new_frihours', 'new_sathours', 'new_sunhours', 'new_totalhours',
        'crd63_timesheetinitiated', 'new_timesheetstatus', 'crd63_weekstartdate', 'new_comments'
    ]

    # Escribir encabezados en la fila 1
    for col, encabezado in enumerate(encabezados, 1):
        ws.cell(row=1, column=col, value=encabezado)

    # Escribir los datos
    for row_idx, (_, fila) in enumerate(df_final.iterrows(), 2):
        for col_idx, encabezado in enumerate(encabezados, 1):
            valor = fila[encabezado]
            celda = ws.cell(row=row_idx, column=col_idx, value=valor)

            # Aplicar formatos específicos según la columna
            if col_idx == 15:  # Columna O (crd63_weekstartdate)
                # Convertir a formato nativo de Excel y aplicar formato
                celda.value = to_excel(valor)  # Convierte datetime a número de Excel
                celda.number_format = 'M/D/YY'  # Formato de fecha corta
            elif col_idx in [5, 6, 7, 8, 9, 10, 11, 12]:  # Columnas E-L (horas)
                # Solo escribir si el valor no es 0, sino dejar vacío
                if valor == 0:
                    celda.value = None
                else:
                    celda.value = float(valor)
                    celda.number_format = 'General'
            elif col_idx == 13:  # Columna M (crd63_timesheetinitiated)
                celda.value = True
                celda.number_format = 'General'  # Excel mostrará como TRUE
            else:  # Resto de columnas (texto)
                celda.number_format = 'General'

    # Crear tabla de Excel
    if len(df_final) > 0:
        # Definir el rango de la tabla (desde A1 hasta la última celda con datos)
        ultima_fila = len(df_final) + 1  # +1 por los encabezados
        ultima_columna = len(encabezados)
        ultima_columna_letra = get_column_letter(ultima_columna)

        rango_tabla = f"A1:{ultima_columna_letra}{ultima_fila}"

        # Crear la tabla con el mismo nombre que el archivo original
        tabla = Table(displayName=NameTableSheet, ref=rango_tabla)

        # Aplicar estilo de tabla (estilo básico de Excel)
        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tabla.tableStyleInfo = estilo

        # Agregar la tabla a la hoja
        ws.add_table(tabla)

    # Ajustar anchos de columna (replicar los del archivo original)
    anchos_columnas = [
        10.08, 19.18, 19.27, 16.63, 15.00, 14.09, 14.73, 15.63,
        13.09, 14.00, 14.36, 15.18, 23.09, 20.18, 19.91, 15.54
    ]

    for col, ancho in enumerate(anchos_columnas, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Guardar el archivo
    wb.save(ruta_guardado)

    print(f"Excel file created: {ruta_guardado}")
    print(f"Sheet name: {ws.title}")
    print(f"Table name: {NameTableSheet}")
    print(f"Rows processed: {len(df_final)}")

    return ruta_guardado


# =============================================================================
# GESTIÓN DE BASE DE DATOS
# =============================================================================
# =============================================================================
# DESCARGA DE BASE DE DATOS DEL N4W - BOX
# =============================================================================
def Download_DataBase_N4W_Box(url_box, salida):
    # Convertir URL de preview a URL de descarga directa
    url_descarga = url_box.replace('/s/', '/shared/static/')

    # Descargar el archivo
    response = requests.get(url_descarga)
    response.raise_for_status()  # Verificar que la descarga fue exitosa

    # Guardar el archivo
    with open(salida, 'wb') as archivo:
        archivo.write(response.content)

    print(f"File successfully downloaded to: {salida}")


# =============================================================================
# GESTIÓN DE CATEGORÍAS DE OUTLOOK
# =============================================================================

def show_progress_window(max_value):
    """Muestra ventana de progreso global."""
    global progress_window, progress_bar
    if app_instance:
        app_instance.show_progress_window(max_value)
        progress_window = app_instance.progress_window
        progress_bar = app_instance.progress_bar


def hide_progress_window():
    """Oculta ventana de progreso global."""
    global progress_window, progress_bar
    if app_instance:
        app_instance.hide_progress_window()
        progress_window = None
        progress_bar = None


def update_categories(filepath, url_box="https://tnc.box.com/s/6y6iswltvf26pxrk3rt1e5s2i7xfo7k4"):
    """
    Actualiza las categorías en Outlook basándose en el archivo Excel.

    Args:
        filepath (str): Ruta al archivo Excel con las categorías
    """
    outlook = None
    com_initialized = False

    try:
        # Inicializar COM para este thread
        pythoncom.CoInitialize()
        com_initialized = True
        print("COM initialized for thread")

        ProjectPath = os.path.dirname(filepath)

        # Ruta de salida de archivo de códigos del N4W
        PathDB_N4W_Box = os.path.join(ProjectPath, "N4W_Task_Details.xlsx")

        # Descarga archivo de códigos del N4W
        Download_DataBase_N4W_Box(url_box, PathDB_N4W_Box)

        # Actualizar base de datos (Ojo, apago este modulo para workday)
        Update_DataBase_With_BoxFile(filepath, PathDB_N4W_Box)

        # Leer y validar datos
        df = readDataBase(filepath)
        df = df.dropna(subset=['Code']).fillna(0)

        required_columns = ['Category', 'Include']
        for column in required_columns:
            if column not in df.columns:
                raise ValueError(f"The Excel file must contain a column '{column}'.")

        # Mostrar progreso
        total_items = len(df)
        show_progress_window(total_items)

        # Conectar con Outlook
        outlook = win32com.client.Dispatch("Outlook.Application")
        categories = outlook.Session.Categories

        # Obtener lista de categorías existentes
        existing_categories = []
        for i in range(1, categories.Count + 1):
            existing_categories.append(categories.Item(i).Name)

        # Procesar categorías
        for i, row in df.iterrows():
            category_name = row['Category']
            include = row['Include']
            color_index = row.get('ColorIndex', i % 25 + 1)

            if include == 1:
                if category_name not in existing_categories:
                    categories.Add(category_name, color_index)
            elif include == 0:
                if category_name in existing_categories:
                    categories.Remove(category_name)

            time.sleep(0.25)

            # Pump COM messages cada 10 iteraciones para evitar desconexiones
            if i % 10 == 0:
                pythoncom.PumpWaitingMessages()

            # Actualizar progreso
            if progress_bar:
                progress_bar['value'] = i + 1
                progress_window.update_idletasks()

        hide_progress_window()
        messagebox.showinfo("Completed", "Category update completed.")

        # Habilitar botones al completar exitosamente
        if app_instance:
            app_instance.enable_all_action_buttons()

    except Exception as e:
        hide_progress_window()
        messagebox.showerror("Error", f"Error updating categories: {e}")

        # Habilitar botones incluso si hay error
        if app_instance:
            app_instance.enable_all_action_buttons()

    finally:
        # Liberar objeto Outlook COM
        try:
            if outlook is not None:
                outlook = None
                print("Outlook COM instance released")
        except Exception as e:
            print(f"Warning: Error releasing Outlook COM: {e}")

        # Liberar COM del thread
        if com_initialized:
            try:
                pythoncom.CoUninitialize()
                print("COM uninitialized for thread")
            except Exception as e:
                print(f"Warning: Error uninitializing COM: {e}")


def run_update_categories(filepath):
    """Ejecuta la actualización de categorías en hilo separado."""
    threading.Thread(target=update_categories, args=(filepath,), daemon=True).start()


# =============================================================================
# EXTRACCIÓN DE DATOS DE CALENDARIO
# =============================================================================

def get_calendar(start_date, end_date, buffer_start=25, buffer_end=25):
    """
    Extrae reuniones del calendario de Outlook en un rango de fechas.

    Args:
        start_date (str): Fecha inicio en formato 'YYYY-MM-DD'
        end_date (str): Fecha fin en formato 'YYYY-MM-DD'
        buffer_start (int): Días adicionales antes del inicio
        buffer_end (int): Días adicionales después del fin

    Returns:
        pd.DataFrame: DataFrame con reuniones extraídas
    """
    outlook = None
    namespace = None

    try:
        # Conectar con Outlook
        outlook_app = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook_app.GetNamespace("MAPI")
        calendar = namespace.GetDefaultFolder(9)

        # Configurar zona horaria
        local_tz = get_localzone()
        start_date = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=local_tz)
        end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=local_tz)

        # Obtener elementos con buffer
        items = calendar.Items
        items.IncludeRecurrences = True
        items.Sort("[Start]")

        start_buffer = start_date - timedelta(days=buffer_start)
        end_buffer = end_date + timedelta(days=buffer_end)

        # Filtrar por fechas
        start_str = start_buffer.strftime('%m/%d/%Y %H:%M')
        end_str = end_buffer.strftime('%m/%d/%Y %H:%M')
        restriction = f"[Start] >= '{start_str}' AND [End] <= '{end_str}'"
        restricted_items = items.Restrict(restriction)

        # Extraer reuniones
        meetings = []
        for item in restricted_items:
            try:
                meeting_start = item.Start
                meeting_end = item.End

                if remove_timezone(start_date) <= remove_timezone(meeting_start) <= remove_timezone(end_date):
                    category = item.Categories if item.Categories else "Sin Category"
                    meeting_date = meeting_start.date()
                    duration = (meeting_end - meeting_start).total_seconds() / 3600

                    meetings.append({
                        'Date': meeting_date,
                        'Category': category,
                        'Hours': duration
                    })
            except AttributeError:
                continue

        return pd.DataFrame(meetings)

    finally:
        # Liberar objetos Outlook COM
        try:
            if namespace is not None:
                namespace = None
            if outlook_app is not None:
                outlook_app = None
            print("Outlook COM instances released")
        except Exception as e:
            print(f"Warning: Error releasing Outlook COM: {e}")


def calculate_workdays(year, month):
    """Calcula días laborables en un mes (excluyendo fines de semana)."""
    _, total_days = calendar.monthrange(year, month)
    workdays = sum(1 for day in range(1, total_days + 1)
                   if datetime(year, month, day).weekday() < 5)
    return workdays


def process_category(category):
    """
    Procesa y clasifica categorías de reuniones.

    Args:
        category (str): Categoría original de la reunión

    Returns:
        tuple: (tipo_ganancia, categoría_limpia)
    """
    keywords = [
        "REGULAR", "LWOP", "MATERNITY", "ADMIN LEAVE", "PARENTAL LEAVE",
        "Compensation", "FURLOUGH", "PUBLIC HOLIDAY", "Medical Leave",
        "Personal Leave Day", "SICK", "VACATION"
    ]

    # Buscar palabra clave
    found_keyword = next(
        (keyword for keyword in keywords if re.search(keyword, category, flags=re.IGNORECASE)),
        "REGULAR"
    )

    # Limpiar categoría
    if found_keyword != "REGULAR":
        category = re.sub(found_keyword, "", category, flags=re.IGNORECASE)

    category = category.replace(",", "").replace(";", "").strip()

    return found_keyword, category


# =============================================================================
# GENERACIÓN DE REPORTES
# =============================================================================

def generate_report(start_date, end_date, database_name, url_box="https://tnc.box.com/s/6y6iswltvf26pxrk3rt1e5s2i7xfo7k4"):
    """
    Genera reporte principal combinando datos de calendario y base de datos.

    Args:
        start_date (datetime): Fecha de inicio
        end_date (datetime): Fecha de fin
        database_name (str): Ruta a la base de datos
    """
    try:

        # Ruta del proyecto
        ProjectPath = os.path.dirname(database_name)

        # Ruta de salida de archivo de códigos del N4W
        PathDB_N4W_Box = os.path.join(ProjectPath, "N4W_Task_Details.xlsx")

        # Descarga archivo de códigos del N4W
        Download_DataBase_N4W_Box(url_box, PathDB_N4W_Box)

        # Actualizar base de datos
        Update_DataBase_With_BoxFile(database_name, PathDB_N4W_Box)

        local_tz = get_localzone()

        # Validar fechas
        if start_date > end_date:
            messagebox.showerror("Error", "The start date cannot be later than the end date.")

            # Habilitar botones cuando hay error de validación
            if app_instance:
                app_instance.enable_all_action_buttons()
            return

        end_date = end_date + timedelta(days=1)

        # Intentar múltiples configuraciones de buffer para obtener datos
        buffer_configs = [13, 2, 3, 5, 7, 11, 17, 19, 23, 29, 31]
        results = pd.DataFrame()

        for buffer1 in buffer_configs:
            for buffer2 in buffer_configs:
                results = get_calendar(
                    start_date.strftime('%Y-%m-%d'),
                    end_date.strftime('%Y-%m-%d'),
                    buffer1, buffer2
                )
                if len(results.columns) != 0:
                    break
            if len(results.columns) != 0:
                break

        # # Procesar categorías
        # results[['Earning', 'Category']] = results['Category'].apply(
        #     lambda x: pd.Series(process_category(x))
        # )

        # Agregar y reorganizar datos
        tmp = results.groupby(by=['Date', 'Category'], as_index=False)['Hours'].sum()

        # Redondear horas a precisión de 0.25
        tmp['Hours'] = tmp['Hours'].apply(lambda x: round(x * 4) / 4)
        tmp = tmp.pivot(index=['Category'], columns='Date', values='Hours').fillna(0)
        # tmp = tmp.reset_index(level='Earning')

        # Crear reporte con fechas completas
        report = pd.DataFrame(columns=pd.date_range(start_date, end_date, freq='D'))
        tmp.columns = pd.to_datetime(tmp.columns, errors='coerce')
        report.columns = pd.to_datetime(report.columns, errors='coerce')
        report = pd.concat([report, tmp], axis=0).fillna(0)

        # Formatear columnas de fechas
        report.columns = report.columns.map(
            lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if isinstance(x, pd.Timestamp) else x
        )

        report.index = [texto.split('|')[0].strip() for texto in report.index.values]

        # Combinar con códigos N4W
        n4w_codes = readDataBase(database_name)
        n4w_codes = n4w_codes.dropna(subset=['Code']).fillna(0).replace('XXXXXX', 0)
        # n4w_codes['Activity ID'] = n4w_codes['Activity ID'].astype(int)
        # n4w_codes['Project ID'] = n4w_codes['Project ID'].astype(str)
        # n4w_codes['Award ID'] = n4w_codes['Award ID'].astype(str)
        n4w_codes = n4w_codes.set_index(['Code'])

        # Crear archivo final
        output_dir = os.path.dirname(database_name)
        value = pd.merge(n4w_codes, report, left_index=True, right_index=True)
        value = value.drop(columns=['Description', 'Category', 'Include'])
        # value.columns = [str(col) if pd.notnull(col) else 'Earning' for col in value.columns]

        # # Reorganizar columnas
        # cols = value.columns.tolist()
        # earning_col = 'Earning'
        # cols.insert(3, cols.pop(cols.index(earning_col)))
        # value = value[cols]
        #
        # # Mapear códigos de ganancia
        # earning_mapping = {
        #     'REGULAR': '1', 'LWOP': '17', 'MATERNITY': '301', 'ADMIN LEAVE': '6',
        #     'PARENTAL LEAVE': '69', 'Compensation': 'C', 'FURLOUGH': 'FRL',
        #     'PUBLIC HOLIDAY': 'H', 'Medical Leave': 'ML', 'Personal Leave Day': 'PLD',
        #     'SICK': 'S', 'VACATION': 'V'
        # }
        # value['Earning'] = value['Earning'].map(earning_mapping)

        # Eliminar última columna (día adicional)
        value = value.drop(columns=value.columns[-1])

        # Guardar archivos
        create_folder(output_dir)
        results.to_excel(os.path.join(output_dir, '01-Report.xlsx'))
        value.to_csv(os.path.join(output_dir, '02-Timesheet.csv'), index_label='Code')

        messagebox.showinfo("Completed", "Process successfully completed.")

        # Habilitar botones al completar exitosamente
        if app_instance:
            app_instance.enable_all_action_buttons()

    except Exception as e:
        messagebox.showerror("General Error", f"Unexpected error: {e}")
        traceback.print_exc()

        # Habilitar botones incluso si hay error
        if app_instance:
            app_instance.enable_all_action_buttons()


# =============================================================================
# AUTOMATIZACIÓN WEB - DELTEK
# =============================================================================
def fill_deltek(position, login_id, password, database_name, prorate=False,
                url_box="https://tnc.box.com/s/6y6iswltvf26pxrk3rt1e5s2i7xfo7k4"):
    """
    Automatiza el llenado de formularios en Deltek usando Selenium WebDriver.

    Args:
        position (int): Posición inicial en la tabla de Deltek
        login_id (str): ID de usuario para login en Deltek
        password (str): Contraseña para login en Deltek
        database_name (str): Ruta al archivo de base de datos de proyectos
        prorate (bool): Si aplicar redistribución de horas de proyectos virtuales
        url_box (str): URL del archivo de códigos N4W en Box
    """
    try:
        # # Ruta del proyecto
        ProjectPath = os.path.dirname(database_name)
        #
        # # Ruta de salida de archivo de códigos del N4W
        PathDB_N4W_Box = os.path.join(ProjectPath, "N4W_Task_Details.xlsx")
        #
        # # Descarga archivo de códigos del N4W
        # Download_DataBase_N4W_Box(url_box, PathDB_N4W_Box)
        #
        # # Actualizar base de datos
        # Update_DataBase_With_BoxFile(database_name, PathDB_N4W_Box)

        FileTimeDeltek = os.path.join(ProjectPath, '02-Timesheet.csv')
        if prorate:
            # Example paths - adjust as needed
            output_file = os.path.join(ProjectPath, '03-Timesheet_Prorate.csv')

            # Run redistribution
            redistribute_hours_by_earning(FileTimeDeltek, PathDB_N4W_Box, output_file, database_name)

            # Show comparison window and get user confirmation
            user_approved = show_prorate_comparison_window(
                os.path.join(ProjectPath, '02-Timesheet.csv'),
                output_file,
                database_name
            )

            if not user_approved:
                print("Process cancelled by user after prorate comparison.")
                messagebox.showinfo("Cancelled", "Deltek process cancelled by user.")

                # Habilitar botones cuando el usuario cancela
                if app_instance:
                    app_instance.enable_all_action_buttons()
                return

            FileTimeDeltek = output_file

        # Configuración
        domain = 'TNC.ORG'
        chrome_path = r'chromedriver.exe'

        # Leer datos procesados
        value = pd.read_csv(FileTimeDeltek, index_col=0)
        value = value.groupby(['Project ID', 'Activity ID', 'Award ID', 'Earning'], as_index=False).sum()

        deltek_data = value[['Project ID', 'Activity ID', 'Award ID', 'Earning']]
        value = value.drop(columns=['Project ID', 'Activity ID', 'Award ID', 'Earning'])
        value[np.isnan(value)] = 0
        value.columns = pd.to_datetime(value.columns)

        # Configurar Chrome
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_experimental_option("detach", True)

        try:
            service = Service(executable_path=chrome_path)
            driver = webdriver.Chrome(service=service, options=chrome_options)
        except:
            driver = webdriver.Chrome(chrome_path, chrome_options=chrome_options)

        # Navegar a Deltek
        driver.get("https://tnc.hostedaccess.com/DeltekTC/TimeCollection.msv")
        wait_time = 10

        # Login
        WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input#uid'))
        ).send_keys(login_id)

        WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input#passField'))
        ).send_keys(password)

        WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input#dom'))
        ).send_keys(domain)

        WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input#loginButton'))
        ).click()

        # Limpiar tabla existente
        driver.switch_to.frame(1)
        WebDriverWait(driver, wait_time).until(
            EC.presence_of_element_located((By.ID, "allRowSelector"))
        ).click()
        WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.ID, "deleteLine"))
        ).click()
        time.sleep(0.5)

        # Pausa para verificación de mes
        messagebox.showinfo(
            "Month Verification",
            "IMPORTANT: Deltek automatically changes to the current month.\n\n"
            "If you need to fill a timesheet for a PREVIOUS month:\n"
            "1. Manually navigate to the correct month in Deltek\n"
            "2. Click OK to continue with automatic filling\n\n"
            "If you are filling the CURRENT month, just click OK."
        )

        # Script para centrar elemento manejando scrolls verticales y horizontales de Deltek
        scroll_into_view_script = """
            var element = document.getElementById(arguments[0]);
            if (element) {
                // Primero usar scrollIntoView nativo para posicionamiento general
                element.scrollIntoView({behavior: 'instant', block: 'center', inline: 'center'});

                // Ajustar scrollers personalizados de Deltek si existen
                var vertScroller = document.getElementById('vertScroller');
                var hrsScroller = document.getElementById('hrsScroller');
                var udtScroller = document.getElementById('udtScroller');

                // Ajustar scroll vertical si existe
                if (vertScroller) {
                    var elementTop = element.offsetTop;
                    var scrollerHeight = vertScroller.offsetHeight;
                    var vertScrollerContent = document.getElementById('vertScrollerContent');

                    if (vertScrollerContent) {
                        var maxScroll = vertScrollerContent.offsetHeight - scrollerHeight;
                        var targetScroll = elementTop - (scrollerHeight / 2);
                        targetScroll = Math.max(0, Math.min(targetScroll, maxScroll));
                        vertScroller.scrollTop = targetScroll;
                    }
                }

                // Ajustar scroll horizontal dependiendo del tipo de elemento
                var elementId = element.id;

                // Si es celda de horas (hrs*), usar hrsScroller
                if (elementId.startsWith('hrs') && hrsScroller) {
                    var elementLeft = element.offsetLeft;
                    var scrollerWidth = hrsScroller.offsetWidth;
                    var hrsScrollerContent = document.getElementById('hrsScrollerContent');

                    if (hrsScrollerContent) {
                        var maxScrollH = hrsScrollerContent.offsetWidth - scrollerWidth;
                        var targetScrollH = elementLeft - (scrollerWidth / 2);
                        targetScrollH = Math.max(0, Math.min(targetScrollH, maxScrollH));
                        hrsScroller.scrollLeft = targetScrollH;
                    }
                }
                // Si es celda de datos (udt*), usar udtScroller
                else if (elementId.startsWith('udt') && udtScroller) {
                    var elementLeft = element.offsetLeft;
                    var scrollerWidth = udtScroller.offsetWidth;
                    var udtScrollerContent = document.getElementById('udtScrollerContent');

                    if (udtScrollerContent) {
                        var maxScrollH = udtScrollerContent.offsetWidth - scrollerWidth;
                        var targetScrollH = elementLeft - (scrollerWidth / 2);
                        targetScrollH = Math.max(0, Math.min(targetScrollH, maxScrollH));
                        udtScroller.scrollLeft = targetScrollH;
                    }
                }

                return true;
            }
            return false;
        """

        # Script para horas - scroll horizontal nativo + scroll vertical manual
        scroll_hrs_script = """
            var element = document.getElementById(arguments[0]);
            if (element) {
                // ScrollIntoView nativo para posicionamiento base (maneja horizontal)
                element.scrollIntoView({behavior: 'instant', block: 'center', inline: 'center'});

                // Ajuste MANUAL de vertScroller (necesario para Deltek)
                var vertScroller = document.getElementById('vertScroller');
                if (vertScroller) {
                    var elementTop = element.offsetTop;
                    var scrollerHeight = vertScroller.offsetHeight;
                    var vertScrollerContent = document.getElementById('vertScrollerContent');

                    if (vertScrollerContent) {
                        var maxScroll = vertScrollerContent.offsetHeight - scrollerHeight;
                        var targetScroll = elementTop - (scrollerHeight / 2);
                        targetScroll = Math.max(0, Math.min(targetScroll, maxScroll));
                        vertScroller.scrollTop = targetScroll;
                    }
                }

                // NO tocar hrsScroller manualmente - dejar que scrollIntoView lo maneje

                return true;
            }
            return false;
        """

        # Script para ajustar SOLO scroll vertical (para filas individuales)
        scroll_vertical_only_script = """
            var element = document.getElementById(arguments[0]);
            if (element) {
                var vertScroller = document.getElementById('vertScroller');

                // Si existe vertScroller, usarlo
                if (vertScroller) {
                    var elementTop = element.offsetTop;
                    var scrollerHeight = vertScroller.offsetHeight;
                    var vertScrollerContent = document.getElementById('vertScrollerContent');

                    if (vertScrollerContent) {
                        var maxScroll = vertScrollerContent.offsetHeight - scrollerHeight;
                        var targetScroll = elementTop - (scrollerHeight / 2);
                        targetScroll = Math.max(0, Math.min(targetScroll, maxScroll));
                        vertScroller.scrollTop = targetScroll;
                    }
                }
                // Si no existe vertScroller, usar scrollIntoView nativo solo para vertical
                // IMPORTANTE: inline: 'nearest' previene que se mueva el scroll horizontal
                else {
                    element.scrollIntoView({behavior: 'instant', block: 'nearest', inline: 'nearest'});
                }
                return true;
            }
            return false;
        """


        # Llenar datos del proyecto
        for i in range(len(deltek_data)):
            # Project ID
            element_id = f"udt{i + position}_1"
            element = WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, element_id))
            )
            driver.execute_script(scroll_into_view_script, element_id)
            time.sleep(0.1)
            element.click()
            WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, "editor"))
            ).send_keys(deltek_data["Project ID"].iloc[i])

            # Award ID - campo 3 (skip) y campo 4 (llenar)
            element_id = f"udt{i + position}_4"
            element = WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, element_id))
            )
            driver.execute_script(scroll_into_view_script, element_id)
            time.sleep(0.1)
            element.click()
            editor = WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, "editor"))
            )
            editor.clear()
            editor.send_keys(str(deltek_data["Award ID"].iloc[i]))

            # Activity ID
            element_id = f"udt{i + position}_5"
            element = WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, element_id))
            )
            driver.execute_script(scroll_into_view_script, element_id)
            time.sleep(0.1)
            element.click()
            editor = WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, "editor"))
            )
            editor.clear()
            editor.send_keys(str(deltek_data["Activity ID"].iloc[i]))

            # Earning Code
            element_id = f"udt{i + position}_6"
            element = WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, element_id))
            )
            driver.execute_script(scroll_into_view_script, element_id)
            time.sleep(0.1)
            element.click()
            editor = WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.ID, "editor"))
            )
            editor.clear()
            editor.send_keys(str(deltek_data["Earning"].iloc[i]))

        # Script para cerrar el editor explícitamente
        close_editor_script = """
            var editor = document.getElementById('editor');
            if (editor) {
                editor.blur();
                // Forzar que Deltek procese el cierre
                var event = new Event('blur', { bubbles: true });
                editor.dispatchEvent(event);
            }
        """

        # Llenar horas - scroll horizontal solo una vez por columna
        for j in range(value.shape[1]):
            for i in range(len(deltek_data)):
                element_id = f"hrs{i + position}_{j}"
                element = WebDriverWait(driver, wait_time).until(
                    EC.presence_of_element_located((By.ID, element_id))
                )

                # Solo hacer scroll en la PRIMERA celda de cada columna
                if i == 0:
                    # Primera celda: scroll completo usando script específico para horas
                    driver.execute_script(scroll_hrs_script, element_id)
                    time.sleep(0.2)

                # Para todas las celdas (incluyendo la primera después del scroll)
                element.click()
                editor = WebDriverWait(driver, wait_time).until(
                    EC.presence_of_element_located((By.ID, "editor"))
                )
                editor.clear()
                editor.send_keys(str(value.iloc[i, j]))

                # Cerrar el editor SOLO si no es la última fila de la columna
                # (evita que se devuelva al inicio al cambiar de columna)
                if i < len(deltek_data) - 1:
                    driver.execute_script(close_editor_script)
                    time.sleep(0.05)
                else:
                    # Última fila: solo esperar un momento sin cerrar el editor
                    # El scroll de la siguiente columna lo cerrará automáticamente
                    time.sleep(0.05)

        print("Deltek process completed")
        messagebox.showinfo("Completed", "Deltek process successfully completed.")

        # Habilitar botones al completar exitosamente
        if app_instance:
            app_instance.enable_all_action_buttons()

    except Exception as e:
        messagebox.showerror("Error General", f"Error inesperado: {e}")
        traceback.print_exc()

        # Habilitar botones incluso si hay error
        if app_instance:
            app_instance.enable_all_action_buttons()

# =============================================================================
# WORKDAY FILE
# =============================================================================
def Create_Workday_File(prorate, database_name):
    try:
        # Ruta del proyecto
        ProjectPath = os.path.dirname(database_name)

        # Ruta de salida de archivo de códigos del N4W
        PathDB_N4W_Box = os.path.join(ProjectPath, "N4W_Task_Details.xlsx")

        FileTimeDeltek = os.path.join(ProjectPath, '02-Timesheet.csv')
        if prorate:
            # Example paths - adjust as needed
            output_file = os.path.join(ProjectPath, '03-Timesheet_Prorate.csv')

            # Run redistribution
            redistribute_hours_by_earning(FileTimeDeltek, PathDB_N4W_Box, output_file, database_name)

            # Show comparison window and get user confirmation
            user_approved = show_prorate_comparison_window(
                os.path.join(ProjectPath, '02-Timesheet.csv'),
                output_file,
                database_name
            )

            if not user_approved:
                print("Process cancelled by user after prorate comparison.")
                messagebox.showinfo("Cancelled", "Deltek process cancelled by user.")

                # Habilitar botones cuando el usuario cancela
                if app_instance:
                    app_instance.enable_all_action_buttons()
                return

            FileTimeDeltek = output_file
            # Habilitar botones al completar exitosamente
            if app_instance:
                app_instance.enable_all_action_buttons()

    except Exception as e:
        messagebox.showerror("Error General", f"Error inesperado: {e}")
        traceback.print_exc()

        # Habilitar botones incluso si hay error
        if app_instance:
            app_instance.enable_all_action_buttons()

# =============================================================================
# AUTOMATIZACIÓN WEB - N4W FACILITY
# =============================================================================
def validate_complete_weeks(start_date, end_date):
    """
    Valida que las fechas correspondan a semanas completas (lunes a domingo).
    
    Args:
        start_date (datetime): Fecha de inicio
        end_date (datetime): Fecha de fin
        
    Returns:
        tuple: (is_valid, error_message)
    """
    # Verificar que start_date sea lunes (weekday() = 0)
    if start_date.weekday() != 0:
        return False, f"Start date must be a Monday. Selected date is a {start_date.strftime('%A')}."
    
    # Verificar que end_date sea domingo (weekday() = 6)  
    if end_date.weekday() != 6:
        return False, f"End date must be a Sunday. Selected date is a {end_date.strftime('%A')}."
    
    # Verificar que sea exactamente semanas completas
    days_difference = (end_date - start_date).days + 1
    if days_difference % 7 != 0:
        return False, f"Date range must be complete weeks. Current range is {days_difference} days."
    
    return True, ""


def validate_deltek_file_weeks(deltek_csv_path):
    """
    Valida que las fechas en el archivo 02-Timesheet.csv correspondan a semanas completas.
    
    Args:
        deltek_csv_path (str): Ruta al archivo 02-Timesheet.csv
        
    Returns:
        tuple: (is_valid, error_message, file_start_date, file_end_date)
    """
    try:
        # Leer el archivo CSV
        df = pd.read_csv(deltek_csv_path)
        
        # Identificar columnas de fechas (pueden tener timestamp)
        date_columns = get_date_columns(df)
        
        if not date_columns:
            return False, "No date columns found in 02-Timesheet.csv file.", None, None
        
        # Convertir nombres de columnas a fechas y ordenar
        date_objects = []
        for col in date_columns:
            try:
                # Intentar primero formato con timestamp
                if ' ' in col:
                    date_str = col.split(' ')[0]  # Tomar solo la parte de fecha
                else:
                    date_str = col
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                date_objects.append(date_obj)
            except ValueError:
                continue
        
        if not date_objects:
            return False, "No valid date columns found in 02-Timesheet.csv file.", None, None
        
        date_objects.sort()
        file_start_date = date_objects[0]
        file_end_date = date_objects[-1]
        
        # Validar semanas completas en el archivo
        is_valid, error_msg = validate_complete_weeks(file_start_date, file_end_date)
        
        if not is_valid:
            return False, f"02-Timesheet.csv file dates are not complete weeks: {error_msg}", file_start_date, file_end_date
        
        return True, "", file_start_date, file_end_date
        
    except Exception as e:
        return False, f"Error reading 02-Timesheet.csv: {str(e)}", None, None


def parse_filename_dates(filename):
    """
    Extrae las fechas de inicio y fin de un nombre de archivo.
    
    Args:
        filename (str): Nombre del archivo (ej: email_2024-12-02_to_2024-12-08.xlsx)
        
    Returns:
        tuple: (start_date, end_date) o (None, None) si no puede parsear
    """
    try:
        # Buscar patrón: email_YYYY-MM-DD_to_YYYY-MM-DD.xlsx
        import re
        pattern = r'.*_(\d{4}-\d{2}-\d{2})_to_(\d{4}-\d{2}-\d{2})\.xlsx$'
        match = re.match(pattern, filename)
        
        if match:
            start_str, end_str = match.groups()
            start_date = datetime.strptime(start_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_str, '%Y-%m-%d')
            return start_date, end_date
        
        return None, None
    except Exception:
        return None, None


def check_week_overlaps(new_start, new_end, existing_start, existing_end):
    """
    Verifica si dos rangos de fechas se solapan.
    
    Args:
        new_start, new_end: Fechas del nuevo rango
        existing_start, existing_end: Fechas del rango existente
        
    Returns:
        bool: True si se solapan
    """
    # Lógica: Los rangos NO se solapan si:
    # - El nuevo rango termina antes de que empiece el existente (new_end < existing_start), O
    # - El nuevo rango empieza después de que termine el existente (new_start > existing_end)
    # Por lo tanto, SE SOLAPAN si NO se cumple ninguna de estas condiciones
    
    no_overlap = (new_end < existing_start or new_start > existing_end)
    overlap = not no_overlap
    
    print(f"    DEBUG overlap check:")
    print(f"      new_end < existing_start: {new_end} < {existing_start} = {new_end < existing_start}")
    print(f"      new_start > existing_end: {new_start} > {existing_end} = {new_start > existing_end}")
    print(f"      No overlap: {no_overlap}")
    print(f"      Overlap: {overlap}")
    
    return overlap


def find_existing_timesheets_in_onedrive(email):
    """
    Busca archivos de timesheet existentes en OneDrive para un email específico.
    Usa la función resolve_onedrive_target existente para detectar OneDrive correctamente.
    
    Args:
        email (str): Email del usuario
        
    Returns:
        list: Lista de tuplas (filename, start_date, end_date) de archivos encontrados
    """
    try:
        # Usar la función existente para obtener la ruta de OneDrive/Tester_TimeSheet
        tester_folder_path = resolve_onedrive_target("N4WTimeTracking - Science Timesheets")

        if not tester_folder_path.exists():
            print(f"OneDrive Tester_TimeSheet folder not found: {tester_folder_path}")
            return []
        
        existing_files = []
        
        # Buscar archivos .xlsx que empiecen con el email
        pattern = f"{email}_*.xlsx"
        
        for file_path in tester_folder_path.glob(pattern):
            filename = file_path.name
            start_date, end_date = parse_filename_dates(filename)
            
            if start_date and end_date:
                existing_files.append((filename, start_date, end_date))
        
        return existing_files
        
    except Exception as e:
        print(f"Error searching OneDrive files: {e}")
        return []


def validate_no_duplicate_weeks(email, new_start_date, new_end_date):
    """
    Valida que no existan semanas duplicadas en los archivos de OneDrive.
    
    Args:
        email (str): Email del usuario
        new_start_date (datetime.date or datetime.datetime): Fecha de inicio del nuevo reporte
        new_end_date (datetime.date or datetime.datetime): Fecha de fin del nuevo reporte
        
    Returns:
        tuple: (is_valid, error_message, conflicting_files)
    """
    try:
        # Buscar archivos existentes
        existing_files = find_existing_timesheets_in_onedrive(email)
        
        print(f"DEBUG: Found {len(existing_files)} existing files for {email}")
        
        if not existing_files:
            return True, "", []
        
        # Convertir fechas a datetime.date para comparación consistente
        if hasattr(new_start_date, 'date'):
            new_start = new_start_date.date()
        else:
            new_start = new_start_date
            
        if hasattr(new_end_date, 'date'):
            new_end = new_end_date.date()
        else:
            new_end = new_end_date
        
        # Verificar solapamientos
        conflicts = []
        
        for filename, existing_start, existing_end in existing_files:
            # Convertir fechas existentes a datetime.date
            if hasattr(existing_start, 'date'):
                existing_start_date = existing_start.date()
            else:
                existing_start_date = existing_start
                
            if hasattr(existing_end, 'date'):
                existing_end_date = existing_end.date()
            else:
                existing_end_date = existing_end
            
            print(f"DEBUG: Checking {filename}")
            print(f"  New range: {new_start} to {new_end}")
            print(f"  Existing range: {existing_start_date} to {existing_end_date}")
            
            if check_week_overlaps(new_start, new_end, existing_start_date, existing_end_date):
                print(f"  -> OVERLAP DETECTED!")
                conflicts.append({
                    'filename': filename,
                    'start': existing_start_date,
                    'end': existing_end_date
                })
            else:
                print(f"  -> No overlap")
        
        if conflicts:
            # Crear mensaje de error detallado
            conflict_details = []
            for conflict in conflicts:
                week_range = f"{conflict['start'].strftime('%Y-%m-%d')} to {conflict['end'].strftime('%Y-%m-%d')}"
                conflict_details.append(f"• Week {week_range} (file: {conflict['filename']})")
            
            error_msg = (
                f"The following weeks have already been submitted:\n\n"
                + "\n".join(conflict_details) +
                f"\n\nNew report range: {new_start.strftime('%Y-%m-%d')} to {new_end.strftime('%Y-%m-%d')}"
            )
            
            return False, error_msg, conflicts
        
        return True, "", []
        
    except Exception as e:
        print(f"Error validating duplicate weeks: {e}")
        import traceback
        traceback.print_exc()
        return True, "", []  # En caso de error, permitir continuar


def get_outlook_active_email():
    """
    Detecta el correo electrónico de la cuenta activa en Outlook.

    Returns:
        str: Dirección de correo de la cuenta activa, o None si no se puede detectar
    """
    outlook = None
    namespace = None

    try:
        import win32com.client

        # Conectar a Outlook
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")

        # Obtener la cuenta por defecto (primera cuenta configurada)
        # Esto funciona para la mayoría de casos donde hay una cuenta principal
        accounts = namespace.Accounts
        if accounts.Count > 0:
            default_account = accounts.Item(1)  # Primera cuenta (índice 1 en COM)
            return default_account.SmtpAddress

        return None

    except Exception as e:
        print(f"Error detecting Outlook email: {e}")
        return None

    finally:
        # Liberar objetos Outlook COM
        try:
            if namespace is not None:
                namespace = None
            if outlook is not None:
                outlook = None
        except Exception as e:
            print(f"Warning: Error releasing Outlook COM: {e}")


def validate_outlook_email_match(user_email):
    """
    Valida que el correo ingresado por el usuario coincida con el correo activo en Outlook.
    
    Args:
        user_email (str): Correo ingresado por el usuario
        
    Returns:
        tuple: (is_valid, error_message, outlook_email)
    """
    try:
        outlook_email = get_outlook_active_email()
        
        if outlook_email is None:
            return True, "", None  # Si no se puede detectar, permitir continuar
        
        # Normalizar emails para comparación (convertir a minúsculas)
        user_email_normalized = user_email.lower().strip()
        outlook_email_normalized = outlook_email.lower().strip()
        
        if user_email_normalized == outlook_email_normalized:
            return True, "", outlook_email
        else:
            error_msg = (
                f"Email mismatch detected:\n\n"
                f"• User entered: {user_email}\n"
                f"• Outlook active account: {outlook_email}\n\n"
                f"Please ensure you're using the same email address that's configured in Outlook."
            )
            return False, error_msg, outlook_email
            
    except Exception as e:
        print(f"Error validating Outlook email match: {e}")
        return True, "", None  # En caso de error, permitir continuar


def Fill_N4W(LoginID, NameDataBase, start_date, end_date, url_box="https://tnc.box.com/s/6y6iswltvf26pxrk3rt1e5s2i7xfo7k4"):
    """
    Automatiza el llenado de formularios en N4W Facility

    Args:
        LoginID (str): Email de usuario
        NameDataBase (str): Ruta a la base de datos
        start_date (datetime): Fecha de inicio
        end_date (datetime): Fecha de fin
        url_box (str): URL del archivo de códigos N4W en Box
    """
    try:
        ProjectPath = os.path.dirname(NameDataBase)
        #
        # # Ruta de salida de archivo de códigos del N4W
        PathDB_N4W_Box = os.path.join(ProjectPath, "N4W_Task_Details.xlsx")
        #
        # # Descarga archivo de códigos del N4W
        # Download_DataBase_N4W_Box(url_box, PathDB_N4W_Box)
        #
        # # Actualizar base de datos
        # Update_DataBase_With_BoxFile(NameDataBase, PathDB_N4W_Box)

        # Validar que el archivo 02-Timesheet.csv tenga semanas completas
        deltek_csv_path = os.path.join(ProjectPath, "02-Timesheet.csv")
        file_valid, file_error, file_start, file_end = validate_deltek_file_weeks(deltek_csv_path)

        if not file_valid:
            messagebox.showerror(
                "Invalid 02-Timesheet.csv File",
                f"The 02-Timesheet.csv file does not contain complete weeks.\n\n{file_error}\n\n"
                f"Please regenerate the Deltek report with complete weeks (Monday to Sunday)."
            )
            # Habilitar botones antes de salir
            if app_instance:
                app_instance.enable_all_action_buttons()
            return
        
        # Verificar que las fechas seleccionadas coincidan con las del archivo
        # Convertir datetime a date para comparación
        file_start_date = file_start.date() if hasattr(file_start, 'date') else file_start
        file_end_date = file_end.date() if hasattr(file_end, 'date') else file_end
        
        if start_date != file_start_date or end_date != file_end_date:
            messagebox.showerror(
                "Date Mismatch",
                f"Selected dates don't match the 02-Timesheet.csv file dates.\n\n"
                f"Selected: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}\n"
                f"File contains: {file_start_date.strftime('%Y-%m-%d')} to {file_end_date.strftime('%Y-%m-%d')}\n\n"
                f"Please select the same date range as in the Deltek file."
            )
            # Habilitar botones antes de salir
            if app_instance:
                app_instance.enable_all_action_buttons()
            return

        # Validar que el correo ingresado coincida con el correo activo en Outlook
        email_valid, email_error, outlook_email = validate_outlook_email_match(LoginID)

        if not email_valid:
            messagebox.showerror(
                "Email Mismatch",
                f"The email entered does not match the active Outlook account.\n\n{email_error}\n\n"
                f"Please use the correct email address or switch to the appropriate Outlook account."
            )
            # Habilitar botones antes de salir
            if app_instance:
                app_instance.enable_all_action_buttons()
            return

        # Verificar correo y obtiene el nombre del usuario
        info = Lookup_UserName_Outlook(LoginID)

        # Validar que no existan semanas duplicadas en OneDrive
        valid_weeks, duplicate_error, conflicts = validate_no_duplicate_weeks(info["email"], start_date, end_date)

        if not valid_weeks:
            messagebox.showerror(
                "Duplicate Weeks Detected",
                f"Cannot submit timesheet due to duplicate weeks.\n\n{duplicate_error}\n\n"
                f"Please check your OneDrive folder and remove conflicting files, or select different weeks."
            )
            # Habilitar botones antes de salir
            if app_instance:
                app_instance.enable_all_action_buttons()
            return

        # Construir nombre de archivo con rango de fechas
        start_str = start_date.strftime("%Y-%m-%d")
        end_str = end_date.strftime("%Y-%m-%d")
        NameFile = f'{info["email"]}_{start_str}_to_{end_str}.xlsx'
        CreateExcel_N4WFormat(archivo_csv=os.path.join(ProjectPath, "02-Timesheet.csv"),
                              email_empleado=info['email'], nombre_empleado=info['name'],
                              ruta_guardado=os.path.join(ProjectPath, NameFile),
                              archivo_base_datos=PathDB_N4W_Box)

        # Enviar archivo a OneDrive
        # put_file_in_onedrive(
        #     os.path.join(ProjectPath, NameFile),
        #     fr"N4WTimeTracking - Science Timesheets\{NameFile}",
        #     account_hint="The Nature Conservancy",  # o parte del nombre de la empresa
        #     overwrite=True
        # )

        messagebox.showinfo("Completed", "N4W Facility process successfully completed.")

        # Habilitar botones al completar exitosamente
        if app_instance:
            app_instance.enable_all_action_buttons()

    except Exception as e:
        messagebox.showerror("Error General", f"Error inesperado: {e}")
        traceback.print_exc()

        # Habilitar botones incluso si hay error
        if app_instance:
            app_instance.enable_all_action_buttons()


# =============================================================================
# INTERFAZ GRÁFICA - APLICACIÓN PRINCIPAL
# =============================================================================
class TimesheetApp:
    """Aplicación principal de automatización de hojas de tiempo."""

    def __init__(self):
        global app_instance
        app_instance = self

        # Crear ventana principal
        self.app = ctk.CTk()
        self.app.title("Timesheet Autofill Tool")
        self.app.geometry("590x780")
        self.app.configure(fg_color=COLORS['bg_primary'])

        # Variables para barra de progreso
        self.progress_window = None
        self.progress_bar = None

        # Configurar grid
        self.app.grid_columnconfigure(0, weight=1)

        self.create_widgets()

    def create_widgets(self):
        """Crea todos los widgets de la interfaz."""
        # Contenedor principal
        main_container = ctk.CTkScrollableFrame(
            self.app,
            fg_color=COLORS['bg_primary'],
            corner_radius=0
        )
        main_container.grid(row=0, column=0, sticky="nsew", padx=24, pady=20)
        main_container.grid_columnconfigure(0, weight=1)
        self.app.grid_rowconfigure(0, weight=1)

        # Header
        self._create_header(main_container)

        # Módulos
        self._create_module1_categories(main_container)
        self._create_module2_meetings(main_container)
        self._create_module3_deltek(main_container)
        self._create_module4_CodeN4W(main_container)

    def _create_header(self, parent):
        """Crea el header de la aplicación."""
        header_frame = ctk.CTkFrame(parent, fg_color="transparent", corner_radius=0)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 24))
        header_frame.grid_columnconfigure(0, weight=1)

        title_label = ctk.CTkLabel(
            header_frame,
            text="Timesheet Autofill Tool",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=COLORS['text_primary']
        )
        title_label.grid(row=0, column=0, sticky="w")

        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Automate your workflow between Outlook, Deltek and N4W Facility",
            font=ctk.CTkFont(size=14),
            text_color=COLORS['text_secondary']
        )
        subtitle_label.grid(row=1, column=0, sticky="w", pady=(4, 0))

    def _create_module1_categories(self, parent):
        """Crea el módulo 1: Actualizar categorías en Outlook."""
        module1 = self.create_module_frame(parent, 1)

        self.create_module_header(
            module1, "01", "Update Outlook Categories",
            "Load your project database and sync categories with the calendar"
        )

        content1 = ctk.CTkFrame(
            module1,
            fg_color=COLORS['bg_secondary'],
            corner_radius=8,
            border_width=1,
            border_color=COLORS['border']
        )
        content1.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 16))
        content1.grid_columnconfigure(0, weight=1)

        input_frame = ctk.CTkFrame(content1, fg_color="transparent")
        input_frame.grid(row=0, column=0, sticky="ew", padx=16, pady=12)
        input_frame.grid_columnconfigure(0, weight=1)

        self.projects_database = ctk.CTkEntry(
            input_frame,
            height=36,
            font=ctk.CTkFont(size=13),
            fg_color=COLORS['bg_tertiary'],
            border_color=COLORS['border'],
            text_color=COLORS['text_primary'],
            placeholder_text="Project database path..."
        )
        self.projects_database.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        button_frame = ctk.CTkFrame(input_frame, fg_color="transparent")
        button_frame.grid(row=0, column=1)

        self.button_load_database = ctk.CTkButton(
            button_frame,
            text="Browse",
            command=lambda: self.select_file(self.projects_database),
            width=80,
            height=36,
            font=ctk.CTkFont(size=13),
            fg_color=COLORS['bg_tertiary'],
            hover_color=COLORS['border'],
            text_color=COLORS['text_primary'],
            border_width=1,
            border_color=COLORS['border']
        )
        self.button_load_database.grid(row=0, column=0, padx=(0, 8))

        self.button_update_categories = ctk.CTkButton(
            button_frame,
            text="Update Categories",
            command=lambda: self.run_update_categories(self.projects_database.get()),
            width=140,
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLORS['accent'],
            hover_color=COLORS['accent_hover']
        )
        self.button_update_categories.grid(row=0, column=1)

    def _create_module2_meetings(self, parent):
        """Crea el módulo 2: Leer reuniones de Outlook."""
        module2 = self.create_module_frame(parent, 2)

        self.create_module_header(
            module2, "02", "Read Outlook Meetings",
            "Extract meeting data from your calendar for the specified date range"
        )

        content2 = ctk.CTkFrame(
            module2,
            fg_color=COLORS['bg_secondary'],
            corner_radius=8,
            border_width=1,
            border_color=COLORS['border']
        )
        content2.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 16))

        date_frame = ctk.CTkFrame(content2, fg_color="transparent")
        date_frame.grid(row=0, column=0, sticky="ew", padx=16, pady=12)
        date_frame.grid_columnconfigure(2, weight=1)

        # Fecha inicio
        date_container1 = ctk.CTkFrame(
            date_frame,
            fg_color=COLORS['bg_tertiary'],
            corner_radius=6,
            border_width=1,
            border_color=COLORS['border']
        )
        date_container1.grid(row=0, column=0, sticky="w", padx=(0, 8))

        start_label = ctk.CTkLabel(
            date_container1,
            text="Start Date",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['text_secondary']
        )
        start_label.pack(anchor="w", padx=8, pady=(6, 0))

        self.start_date_entry = DateEntry(
            date_container1,
            width=16,
            background=COLORS['accent'],
            foreground='white',
            borderwidth=0,
            date_pattern='yyyy-mm-dd',
            font=('Inter', 11)
        )
        self.start_date_entry.pack(padx=8, pady=(0, 6))

        # Fecha fin
        date_container2 = ctk.CTkFrame(
            date_frame,
            fg_color=COLORS['bg_tertiary'],
            corner_radius=6,
            border_width=1,
            border_color=COLORS['border']
        )
        date_container2.grid(row=0, column=1, sticky="w", padx=(0, 8))

        end_label = ctk.CTkLabel(
            date_container2,
            text="End Date",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['text_secondary']
        )
        end_label.pack(anchor="w", padx=8, pady=(6, 0))

        self.end_date_entry = DateEntry(
            date_container2,
            width=16,
            background=COLORS['accent'],
            foreground='white',
            borderwidth=0,
            date_pattern='yyyy-mm-dd',
            font=('Inter', 11)
        )
        self.end_date_entry.pack(padx=8, pady=(0, 6))

        # Botón leer
        self.read_button = ctk.CTkButton(
            date_frame,
            text="Read Meetings",
            command=lambda: self.generate_report(),
            width=120,
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLORS['accent'],
            hover_color=COLORS['accent_hover']
        )
        self.read_button.grid(row=0, column=2, sticky="e")

    def _create_module3_deltek(self, parent):
        """Crea el módulo 3: Llenar Deltek."""
        module3 = self.create_module_frame(parent, 3)

        self.create_module_header(
            module3, "03", "Timesheet for Workday",
            "Generate a file of your hours worked using Workday coding."
        )

        content3 = ctk.CTkFrame(
            module3,
            fg_color=COLORS['bg_secondary'],
            corner_radius=8,
            border_width=1,
            border_color=COLORS['border']
        )
        content3.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 16))

        deltek_frame = ctk.CTkFrame(content3, fg_color="transparent")
        deltek_frame.grid(row=0, column=0, sticky="ew", padx=16, pady=12)
        deltek_frame.grid_columnconfigure(0, weight=2)
        deltek_frame.grid_columnconfigure(1, weight=2)
        deltek_frame.grid_columnconfigure(2, weight=1)
        deltek_frame.grid_columnconfigure(3, weight=0)  # Para el ícono info
        deltek_frame.grid_columnconfigure(4, weight=0)  # Para Prorate checkbox
        deltek_frame.grid_columnconfigure(5, weight=0)  # Para botón Fill Deltek

        # # ID Usuario
        # self.email_entry_deltek = ctk.CTkEntry(
        #     deltek_frame,
        #     width=70,
        #     height=36,
        #     font=ctk.CTkFont(size=13),
        #     fg_color=COLORS['bg_tertiary'],
        #     border_color=COLORS['border'],
        #     text_color=COLORS['text_primary'],
        #     placeholder_text="User ID"
        # )
        # self.email_entry_deltek.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        #
        # # Contraseña
        # self.password_entry_deltek = ctk.CTkEntry(
        #     deltek_frame,
        #     height=36,
        #     font=ctk.CTkFont(size=13),
        #     fg_color=COLORS['bg_tertiary'],
        #     border_color=COLORS['border'],
        #     text_color=COLORS['text_primary'],
        #     placeholder_text="Password",
        #     show="*"
        # )
        # self.password_entry_deltek.grid(row=0, column=1, sticky="ew", padx=(6, 6))
        #
        # # Posición
        # self.posi_entry_deltek = ctk.CTkEntry(
        #     deltek_frame,
        #     width=36,
        #     height=36,
        #     font=ctk.CTkFont(size=13),
        #     fg_color=COLORS['bg_tertiary'],
        #     border_color=COLORS['border'],
        #     text_color=COLORS['text_primary'],
        #     placeholder_text="Position"
        # )
        # self.posi_entry_deltek.grid(row=0, column=2, sticky="ew", padx=(6, 2))
        # self.posi_entry_deltek.insert(0, "0")
        #
        # # Info icon para Position con tooltip
        # info_icon = ctk.CTkLabel(
        #     deltek_frame,
        #     text="ℹ️",
        #     font=ctk.CTkFont(size=16),
        #     text_color=COLORS['accent'],
        #     cursor="hand2",
        #     width=3
        # )
        # info_icon.grid(row=0, column=3, sticky="w", padx=(1, 0))
        #
        # # Crear tooltip para el ícono de información
        # tooltip_text = (
        #     "Pre-existing Row Position\n\n"
        #     "This number indicates how many rows are already\n"
        #     "occupied by projects in your Deltek timesheet\n"
        #     "that cannot be deleted.\n\n"
        #     "• Use 0 when the timesheet is completely empty\n"
        #     "• Use 1 if there is one existing project\n"
        #     "• Use 2 if there are two existing projects, etc.\n\n"
        # )
        # ToolTip(info_icon, tooltip_text)

        # Prorate checkbox
        self.prorate_checkbox = ctk.CTkCheckBox(
            deltek_frame,
            width=50,
            text="Prorate",
            font=ctk.CTkFont(size=13),
            text_color=COLORS['text_primary']
        )
        self.prorate_checkbox.grid(row=0, column=4, padx=(0, 0))
        self.prorate_checkbox.select()  # Set checked by default

        # Botón llenar Deltek
        self.fill_deltek_button = ctk.CTkButton(
            deltek_frame,
            text="Workday",
            command=lambda: self.fill_Workday(),
            width=90,
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLORS['success'],
            hover_color='#0D6A0D'
        )
        self.fill_deltek_button.grid(row=0, column=5, padx=(6, 0))

    def _create_module4_CodeN4W(self, parent):
        """Crea el módulo 4: Llenar N4W Facility."""
        module4 = self.create_module_frame(parent, 4)

        self.create_module_header(
            module4, "04", "Fill N4W Facility Timesheet",
            "Automate filling your N4W Facility timesheet with meeting data"
        )

        content4 = ctk.CTkFrame(
            module4,
            fg_color=COLORS['bg_secondary'],
            corner_radius=8,
            border_width=1,
            border_color=COLORS['border']
        )
        content4.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 16))

        CodeN4W_frame = ctk.CTkFrame(content4, fg_color="transparent")
        CodeN4W_frame.grid(row=0, column=0, sticky="ew", padx=16, pady=12)
        CodeN4W_frame.grid_columnconfigure(0, weight=2)

        # Email
        self.email_entry_CodeN4W = ctk.CTkEntry(
            CodeN4W_frame,
            height=36,
            width=300,
            font=ctk.CTkFont(size=13),
            fg_color=COLORS['bg_tertiary'],
            border_color=COLORS['border'],
            text_color=COLORS['text_primary'],
            placeholder_text="Email"
        )
        self.email_entry_CodeN4W.grid(row=0, column=0, sticky="ew", padx=(0, 6))

        # Fill N4W Facility button
        self.Fill_N4W_App_button = ctk.CTkButton(
            CodeN4W_frame,
            text="Fill N4W Facility",
            command=lambda: self.Fill_N4W_App(),
            width=150,
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COLORS['success'],
            hover_color='#0D6A0D'
        )
        self.Fill_N4W_App_button.grid(row=0, column=1, padx=(6, 0))

    def create_module_frame(self, parent, row):
        """Crea el frame base para un módulo."""
        module = ctk.CTkFrame(parent, fg_color="transparent", corner_radius=0)
        module.grid(row=row, column=0, sticky="ew", pady=(0, 20))
        module.grid_columnconfigure(0, weight=1)
        return module

    def create_module_header(self, parent, step, title, description):
        """Crea el header de un módulo."""
        header = ctk.CTkFrame(parent, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(0, 8))
        header.grid_columnconfigure(1, weight=1)

        # Indicador de paso
        step_circle = ctk.CTkFrame(
            header,
            width=24,
            height=24,
            fg_color=COLORS['accent'],
            corner_radius=12
        )
        step_circle.grid(row=0, column=0, rowspan=2, padx=(0, 12), sticky="n")
        step_circle.grid_propagate(False)

        step_label = ctk.CTkLabel(
            step_circle,
            text=step,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="white"
        )
        step_label.place(relx=0.5, rely=0.5, anchor="center")

        # Título y descripción
        title_label = ctk.CTkLabel(
            header,
            text=title,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS['text_primary']
        )
        title_label.grid(row=0, column=1, sticky="w")

        desc_label = ctk.CTkLabel(
            header,
            text=description,
            font=ctk.CTkFont(size=12),
            text_color=COLORS['text_secondary']
        )
        desc_label.grid(row=1, column=1, sticky="w", pady=(2, 0))

        return header

    # =========================================================================
    # MÉTODOS DE BARRA DE PROGRESO
    # =========================================================================

    def show_progress_window(self, max_value):
        """Muestra ventana de progreso."""
        self.progress_window = ctk.CTkToplevel(self.app)
        self.progress_window.title("Progreso")
        self.progress_window.geometry("400x120")
        self.progress_window.resizable(False, False)
        self.progress_window.configure(fg_color=COLORS['bg_primary'])

        self.progress_window.transient(self.app)
        self.progress_window.grab_set()

        label = ctk.CTkLabel(
            self.progress_window,
            text="Updating categories, please wait...",
            font=ctk.CTkFont(size=14),
            text_color=COLORS['text_primary']
        )
        label.pack(pady=(20, 10))

        progress_frame = tk.Frame(self.progress_window, bg=COLORS['bg_primary'])
        progress_frame.pack(pady=10)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Custom.Horizontal.TProgressbar",
                        background=COLORS['accent'],
                        troughcolor=COLORS['bg_tertiary'],
                        borderwidth=0,
                        lightcolor=COLORS['accent'],
                        darkcolor=COLORS['accent'])

        self.progress_bar = ttk.Progressbar(
            progress_frame,
            orient="horizontal",
            length=350,
            mode="determinate",
            style="Custom.Horizontal.TProgressbar"
        )
        self.progress_bar.pack()
        self.progress_bar['value'] = 0
        self.progress_bar['maximum'] = max_value

        self.progress_window.update_idletasks()

    def hide_progress_window(self):
        """Oculta ventana de progreso."""
        if self.progress_window:
            self.progress_window.destroy()
            self.progress_window = None
            self.progress_bar = None

    # =========================================================================
    # MÉTODOS DE FUNCIONALIDAD
    # =========================================================================
    def select_file(self, entry_widget):
        """Abre diálogo para seleccionar archivo."""
        filename = filedialog.askopenfilename(
            title="Select Database",
            filetypes=[("Excel file", "*.xlsx"), ("all files", "*.*")]
        )
        if filename:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filename)

    def run_update_categories(self, database_path=None):
        """Ejecuta actualización de categorías."""
        # Deshabilitar botones al inicio
        self.disable_all_action_buttons()

        # Forzar actualización inmediata de la UI
        self.app.update_idletasks()

        try:
            if database_path is None:
                database_path = self.projects_database.get()

            if not database_path:
                messagebox.showerror("Error", "Please select a database.")
                self.enable_all_action_buttons()
                return

            run_update_categories(database_path)

        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error: {e}")
            self.enable_all_action_buttons()

    def generate_report(self):
        """Genera reporte de reuniones."""
        # Deshabilitar botones al inicio
        self.disable_all_action_buttons()

        # Forzar actualización inmediata de la UI
        self.app.update_idletasks()

        try:
            start_date = datetime.strptime(self.start_date_entry.get(), '%Y-%m-%d')
            end_date = datetime.strptime(self.end_date_entry.get(), '%Y-%m-%d')
            database_path = self.projects_database.get()

            if not database_path:
                messagebox.showerror("Error", "Please select a database.")
                self.enable_all_action_buttons()
                return

            generate_report(start_date, end_date, database_path)

        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error: {e}")
            self.enable_all_action_buttons()

    def fill_deltek(self):
        """Llena formularios Deltek."""
        # Deshabilitar botones al inicio
        self.disable_all_action_buttons()

        # Forzar actualización inmediata de la UI
        self.app.update_idletasks()

        try:
            user_id     = self.email_entry_deltek.get()
            password    = self.password_entry_deltek.get()
            position    = self.posi_entry_deltek.get()
            database_path = self.projects_database.get()
            prorate     = self.prorate_checkbox.get()

            if not all([user_id, password, position, database_path]):
                messagebox.showerror("Error", "Please complete all fields.")
                self.enable_all_action_buttons()
                return

            fill_deltek(int(position), user_id, password, database_path, prorate)

        except ValueError:
            messagebox.showerror("Error", "Position must be a number.")
            self.enable_all_action_buttons()
        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error: {e}")
            self.enable_all_action_buttons()

    def fill_Workday(self):
        """Llena formularios Deltek."""
        # Deshabilitar botones al inicio
        self.disable_all_action_buttons()

        # Forzar actualización inmediata de la UI
        self.app.update_idletasks()

        try:
            database_path   = self.projects_database.get()
            prorate         = self.prorate_checkbox.get()

            Create_Workday_File(prorate, database_path)

        except ValueError:
            messagebox.showerror("Error", "Position must be a number.")
            self.enable_all_action_buttons()
        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error: {e}")
            self.enable_all_action_buttons()

    def Fill_N4W_App(self):
        """Llena formularios N4W Facility."""
        # Deshabilitar botones al inicio
        self.disable_all_action_buttons()

        # Forzar actualización inmediata de la UI
        self.app.update_idletasks()

        try:
            email = self.email_entry_CodeN4W.get()
            database_path = self.projects_database.get()

            if not all([email, database_path]):
                messagebox.showerror("Error", "Please complete all fields.")
                self.enable_all_action_buttons()
                return

            # Obtener fechas de los widgets
            start_date = self.start_date_entry.get_date()
            end_date = self.end_date_entry.get_date()

            # Convertir a datetime si son objetos date
            if hasattr(start_date, 'date'):
                start_date = datetime.combine(start_date, datetime.min.time())
            if hasattr(end_date, 'date'):
                end_date = datetime.combine(end_date, datetime.min.time())
            elif isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')

            # Validar semanas completas
            is_valid, error_message = validate_complete_weeks(start_date, end_date)

            if not is_valid:
                messagebox.showerror(
                    "Invalid Date Range",
                    f"Please select complete weeks (Monday to Sunday).\n\n{error_message}\n\n"
                    f"Tips:\n"
                    f"• Start date must be a Monday\n"
                    f"• End date must be a Sunday\n"
                    f"• You can select multiple consecutive weeks"
                )
                self.enable_all_action_buttons()
                return

            Fill_N4W(email, database_path, start_date, end_date)

        except Exception as e:
            # Capturar cualquier error inesperado y habilitar botones
            messagebox.showerror("Error", f"Unexpected error: {e}")
            self.enable_all_action_buttons()

    # =========================================================================
    # MÉTODOS DE CONTROL DE ESTADO DE BOTONES
    # =========================================================================
    def disable_all_action_buttons(self):
        """Deshabilita todos los botones de acción para prevenir clics múltiples."""
        self.button_load_database.configure(state="disabled")
        self.button_update_categories.configure(state="disabled")
        self.read_button.configure(state="disabled")
        self.fill_deltek_button.configure(state="disabled")
        self.Fill_N4W_App_button.configure(state="disabled")

    def enable_all_action_buttons(self):
        """Habilita todos los botones de acción después de completar un proceso."""
        self.button_load_database.configure(state="normal")
        self.button_update_categories.configure(state="normal")
        self.read_button.configure(state="normal")
        self.fill_deltek_button.configure(state="normal")
        self.Fill_N4W_App_button.configure(state="normal")

    def run(self):
        """Inicia la aplicación."""
        self.app.mainloop()


# =============================================================================
# PUNTO DE ENTRADA
# =============================================================================
if __name__ == "__main__":
    # Validar y actualizar ChromeDriver antes de iniciar la GUI
    chromedriver_ready = validate_and_update_chromedriver()

    if not chromedriver_ready:
        print("\nNo se puede iniciar la aplicación sin ChromeDriver compatible.")
        print("Por favor, resuelva el problema e intente nuevamente.")
        input("\nPresione Enter para salir...")
        exit(1)

    # Iniciar la aplicación solo si ChromeDriver está listo
    app = TimesheetApp()
    app.run()
